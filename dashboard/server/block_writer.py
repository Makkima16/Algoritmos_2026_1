# -*- coding: utf-8 -*-
"""
Escritura de resultados de bloque a XLSX, replicando el formato de los exec.py
(KQNodes/exec.py: COLUMNAS_BLOQUE) para que los archivos generados desde la GUI
sean indistinguibles de los de terminal.
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUMNAS_BLOQUE = [
    ("#Prueba", 10),
    ("Alcance o Purview (t+1)", 28),
    ("Mecanismo(t)", 22),
    ("Particion", 48),
    ("Perdida", 16),
    ("Tiempo", 20),
]

_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center")
_ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_TOT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def formatear_tiempo(segundos: float) -> str:
    if segundos < 60:
        return f"{segundos:.4f} s"
    elif segundos < 3600:
        m = int(segundos // 60)
        return f"{m} min {segundos % 60:.2f} s"
    h = int(segundos // 3600)
    resto = segundos % 3600
    return f"{h} h {int(resto // 60)} min {resto % 60:.2f} s"


def escribir_xlsx_bloque(
    ruta_salida: Path,
    filas: list[dict],
    tiempo_total: float,
    tiempo_arranque: float = 0.0,
    estado: str = "",
    candidato: str = "",
) -> None:
    """
    filas: lista en orden original con claves
        {num, alcance, mecanismo, particion, perdida, tiempo_fmt, error}
    tiempo_total: wall-clock del lote completo.
    tiempo_arranque: acumulado de preparación del subsistema (warmup), aparte del
        tiempo de búsqueda; se reporta en su propia fila como en exec.py.
    estado:    estado inicial del lote (cadena binaria), guardado en celda aparte.
    candidato: sistema candidato del lote (cadena binaria), guardado en celda aparte.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    for col_idx, (titulo, ancho) in enumerate(COLUMNAS_BLOQUE, start=1):
        c = ws.cell(row=1, column=col_idx, value=titulo)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = _HDR_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    center = Alignment(horizontal="center", vertical="center")
    part_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    error_font = Font(color="CC0000")

    for i, r in enumerate(filas, start=2):
        relleno = _ALT_FILL if (i % 2 == 0) else None
        ws.cell(row=i, column=1, value=r.get("num")).alignment = center
        ws.cell(row=i, column=2, value=r.get("alcance", "")).alignment = center
        ws.cell(row=i, column=3, value=r.get("mecanismo", "")).alignment = center

        if r.get("error"):
            c = ws.cell(row=i, column=4, value=f"ERROR: {r['error']}")
            c.alignment = part_align
            c.font = error_font
            ws.cell(row=i, column=5, value=None).alignment = center
            ws.cell(row=i, column=6, value=None).alignment = center
        else:
            c = ws.cell(row=i, column=4, value=r.get("particion", ""))
            c.alignment = part_align
            perdida = r.get("perdida")
            ws.cell(row=i, column=5,
                    value=round(perdida, 6) if perdida is not None else None).alignment = center
            ws.cell(row=i, column=6, value=r.get("tiempo_fmt", "")).alignment = center

        for col in range(1, 7):
            if relleno and ws.cell(row=i, column=col).fill.fill_type is None:
                ws.cell(row=i, column=col).fill = relleno
        ws.row_dimensions[i].height = 45

    fila_total = len(filas) + 2
    for col in range(1, 7):
        cc = ws.cell(row=fila_total, column=col)
        cc.fill = _TOT_FILL
        cc.alignment = center
    ws.cell(row=fila_total, column=1, value="Tiempo Total Lote").font = Font(bold=True)
    tc = ws.cell(row=fila_total, column=6, value=formatear_tiempo(tiempo_total))
    tc.font = Font(bold=True)
    ws.row_dimensions[fila_total].height = 22

    # Fila de "arranque del motor" (warmup): preparación del subsistema acumulada,
    # aparte del tiempo de las pruebas. Mismo formato que KQNodes/exec.py.
    fila_warm = fila_total + 1
    warm_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    for col in range(1, 7):
        cc = ws.cell(row=fila_warm, column=col)
        cc.fill = warm_fill
        cc.alignment = center
    ws.cell(row=fila_warm, column=1, value="Arranque motor (warmup)").font = Font(bold=True)
    wc = ws.cell(row=fila_warm, column=6, value=formatear_tiempo(tiempo_arranque))
    wc.font = Font(bold=True)
    ws.row_dimensions[fila_warm].height = 22

    # Filas de metadatos del lote: estado inicial y candidato en celdas aparte
    # (col 1 = etiqueta, col 2 = valor). Su #Prueba no es numérico, así que el
    # frontend las omite de la tabla de pruebas y las muestra como cabecera.
    meta_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for offset, (etiqueta, valor) in enumerate(
        (("Estado inicial", estado), ("Candidato", candidato)), start=1
    ):
        fila_meta = fila_warm + offset
        for col in range(1, 7):
            cc = ws.cell(row=fila_meta, column=col)
            cc.fill = meta_fill
            cc.alignment = center
        ws.cell(row=fila_meta, column=1, value=etiqueta).font = Font(bold=True)
        vc = ws.cell(row=fila_meta, column=2, value=valor or "")
        vc.font = Font(name="Consolas")
        vc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[fila_meta].height = 20

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)
