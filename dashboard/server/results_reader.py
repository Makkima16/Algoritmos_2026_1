# -*- coding: utf-8 -*-
"""
Lectura de datasets, CSV de pruebas y resultados guardados (JSON manual + XLSX bloque),
y utilidades de conversión etiquetas→binario para la ejecución por bloque desde la GUI.
"""

import re
import csv
import json
from datetime import datetime
from pathlib import Path

import openpyxl

SERVER_DIR = Path(__file__).resolve().parent
REPO_ROOT = SERVER_DIR.parents[1]

ALGO_ROOTS = {
    "qnodes": REPO_ROOT / "KQNodes",
    "geomip": REPO_ROOT / "KGeoMIP",
}

# Fuente ÚNICA y compartida de datasets/pruebas para ambos algoritmos
# (la carpeta data/ de la raíz, la misma que usa el suite runner).
DATA_SCRIPTS = REPO_ROOT / "data"

# Carpeta de resultados "manuales" — difiere por algoritmo.
MANUAL_DIRNAME = {"qnodes": "manual", "geomip": "manually"}


def _formatear_tiempo(segundos: float) -> str:
    if segundos < 60:
        return f"{segundos:.4f} s"
    if segundos < 3600:
        m = int(segundos // 60)
        return f"{m} min {segundos % 60:.2f} s"
    h = int(segundos // 3600)
    resto = segundos % 3600
    return f"{h} h {int(resto // 60)} min {resto % 60:.2f} s"


def guardar_resultado_manual(algo: str, req: dict, resp: dict) -> str:
    """Persiste una corrida manual como JSON en results/<manual|manually>/,
    replicando el formato de los exec.py para que sea indistinguible de los
    archivos generados desde la terminal. Devuelve la ruta del archivo escrito.

    req:  petición original (/api/run) con estado, candidato, alcance, mecanismo, k…
    resp: respuesta serializada del worker (perdida_phi, distribuciones, tiempos…).
    """
    root = ALGO_ROOTS[algo]
    results_dir = root / "results" / MANUAL_DIRNAME[algo]
    results_dir.mkdir(parents=True, exist_ok=True)

    stem = Path(req["ruta_tpm"]).stem
    fecha = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    estrategia = resp.get("estrategia", algo)

    # Etiqueta de k en el nombre: KQNodes usa el k solicitado (o "All"); KGeoMIP
    # toma el k efectivo de su estrategia ("...K=3...") o "optimal" si fue libre.
    if algo == "geomip":
        m = re.search(r"K=(\d+)", str(estrategia))
        k_etiqueta = m.group(1) if m else "optimal"
    else:
        k_etiqueta = str(req["k"]) if req.get("k") is not None else "All"

    nombre = f"resultado_{stem}_k{k_etiqueta}_{fecha}.json"
    ruta_salida = results_dir / nombre

    tiempo_total = float(resp.get("tiempo_total_segundos", 0.0))
    tiempo_prep = float(resp.get("tiempo_preparacion_segundos", 0.0))
    # tiempo_total_segundos ya viene normalizado (preparación + búsqueda) para ambos
    # motores; la búsqueda pura es el campo emitido o, como respaldo, total - prep.
    tiempo_busqueda = float(
        resp.get("tiempo_busqueda_segundos", max(0.0, tiempo_total - tiempo_prep))
    )

    res_data = {
        "dataset": Path(req["ruta_tpm"]).name,
        "estado_inicial": req.get("estado", ""),
        "sistema_candidato": req.get("candidato", ""),
        "alcance": req.get("alcance", ""),
        "mecanismo": req.get("mecanismo", ""),
    }
    if algo == "qnodes":
        res_data["k_solicitado"] = req.get("k")
        res_data["permitir_presente_vacio"] = req.get("permitir_presente_vacio", False)
    res_data["estrategia"] = estrategia
    res_data["perdida_phi"] = float(resp.get("perdida_phi", 0.0))
    if algo == "geomip":
        res_data["fundamento_eleccion"] = (
            "Se seleccionó esta k-partición porque presenta la pérdida de información "
            "(EMD) mínima global de todo el abanico evaluado."
        )
    res_data["distribucion_subsistema"] = resp.get("distribucion_subsistema")
    res_data["distribucion_particion"] = resp.get("distribucion_particion")
    res_data["particion"] = resp.get("particion", "")
    res_data["tiempo_busqueda_segundos"] = tiempo_busqueda
    res_data["tiempo_preparacion_segundos"] = tiempo_prep
    res_data["tiempo_total_segundos"] = tiempo_total
    res_data["tiempo_formateado"] = _formatear_tiempo(tiempo_total)

    with open(ruta_salida, "w", encoding="utf-8") as f:
        json.dump(res_data, f, indent=4, ensure_ascii=False)
    return str(ruta_salida)


# ── Etiquetas de nodo ──────────────────────────────────────────────────────

def labels_para(n: int) -> list[str]:
    """Genera etiquetas A..Z, AA, AB... (base-26 biyectiva)."""
    out = []
    for i in range(n):
        s, x = "", i
        while True:
            s = chr(ord("A") + x % 26) + s
            x = x // 26 - 1
            if x < 0:
                break
        out.append(s)
    return out


def letras_a_binario(etiquetas: str, n_nodos: int) -> str:
    """Convierte etiquetas de nodos (ej 'ABCEFG') a máscara binaria de longitud n_nodos."""
    etiquetas = (etiquetas or "").strip().upper()
    validas = labels_para(n_nodos)
    set_validas = set(validas)
    activos: set[str] = set()
    idx = 0
    while idx < len(etiquetas):
        matched = False
        for length in range(min(3, len(etiquetas) - idx), 0, -1):
            lbl = etiquetas[idx:idx + length]
            if lbl in set_validas:
                activos.add(lbl)
                idx += length
                matched = True
                break
        if not matched:
            idx += 1
    return "".join("1" if validas[i] in activos else "0" for i in range(n_nodos))


# ── Datasets (TPMs) ────────────────────────────────────────────────────────

def listar_datasets(algo: str | None = None) -> list[dict]:
    # Datasets compartidos desde data/ (idénticos para ambos algoritmos).
    salida = []
    for tipo, carpeta in (("binaria", "samples_binary"), ("no_binaria", "samples_no_binary")):
        d = DATA_SCRIPTS / carpeta
        if not d.exists():
            continue
        for f in sorted(d.glob("*.csv")):
            m = re.match(r"^N(\d+)", f.stem)
            salida.append({
                "archivo": f.name,
                "ruta": str(f),
                "n": int(m.group(1)) if m else None,
                "tipo": tipo,
            })
    return salida


# ── CSV de pruebas ─────────────────────────────────────────────────────────

def listar_pruebas(algo: str | None = None, n: int | None = None) -> list[dict]:
    # CSVs de pruebas compartidos desde data/Pruebas.
    d = DATA_SCRIPTS / "Pruebas"
    salida = []
    if not d.exists():
        return salida
    for f in sorted(d.glob("*.csv")):
        m = re.search(r"N(\d+)", f.stem)
        f_n = int(m.group(1)) if m else None
        if n is not None and f_n != n:
            continue
        salida.append({"archivo": f.name, "ruta": str(f), "n": f_n})
    return salida


def leer_csv_pruebas(ruta: str) -> list[dict]:
    """Lee un CSV de pruebas y devuelve filas normalizadas {num, alcance, mecanismo}."""
    filas = []
    with open(ruta, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for i, row in enumerate(reader):
            def _g(*nombres):
                for nombre in nombres:
                    for k, v in row.items():
                        if k and k.strip().lower() == nombre.lower():
                            return (v or "").strip()
                return ""
            filas.append({
                "num": _g("#prueba", "prueba", "num", "n") or str(i + 1),
                "alcance": _g("alcance o purview (t+1)", "alcance", "purview", "effect"),
                "mecanismo": _g("mecanismo(t)", "mecanismo", "mechanism", "cause"),
            })
    return filas


# ── Resultados guardados ───────────────────────────────────────────────────

def listar_resultados(algo: str, tipo: str) -> list[dict]:
    """tipo ∈ {manual, block}. Devuelve archivos con metadatos básicos."""
    root = ALGO_ROOTS[algo]
    salida = []
    if tipo == "manual":
        d = root / "results" / MANUAL_DIRNAME[algo]
        if d.exists():
            for f in sorted(d.glob("*.json"), reverse=True):
                estado, candidato = "", ""
                try:
                    with open(f, encoding="utf-8") as fh:
                        jd = json.load(fh)
                    estado = str(jd.get("estado_inicial", "") or "")
                    candidato = str(jd.get("sistema_candidato", "") or "")
                except Exception:
                    pass
                salida.append({
                    "archivo": f.name,
                    "ruta": str(f),
                    "formato": "json",
                    "mtime": f.stat().st_mtime,
                    "estado": estado,
                    "candidato": candidato,
                })
    elif tipo == "block":
        d = root / "results" / "block"
        if d.exists():
            for f in sorted(d.rglob("*.xlsx"), reverse=True):
                meta = _meta_xlsx(str(f))
                salida.append({
                    "archivo": f.name,
                    "ruta": str(f),
                    "subcarpeta": f.parent.name,
                    "formato": "xlsx",
                    "mtime": f.stat().st_mtime,
                    "estado": meta.get("estado", ""),
                    "candidato": meta.get("candidato", ""),
                })
    return salida


# Etiquetas (col 1) de las filas de metadatos del lote escritas por block_writer.py
# y los exec.py. Se extraen a `meta` y se omiten de las filas de datos.
_META_LABELS = {"estado inicial": "estado", "candidato": "candidato"}


def _leer_xlsx(ruta: str) -> dict:
    """Lee un XLSX de bloque: encabezados (fila 1), filas de datos y metadatos
    (estado inicial / candidato, guardados en celdas aparte)."""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not filas:
        return {"columnas": [], "filas": [], "meta": {}}
    columnas = [str(c) if c is not None else "" for c in filas[0]]
    meta: dict = {}
    datos = []
    for fila in filas[1:]:
        etiqueta = str(fila[0]).strip().lower() if fila and fila[0] is not None else ""
        if etiqueta in _META_LABELS:
            valor = fila[1] if len(fila) > 1 else None
            meta[_META_LABELS[etiqueta]] = str(valor) if valor is not None else ""
            continue
        datos.append({columnas[i] if i < len(columnas) else f"col{i}": fila[i]
                      for i in range(len(fila))})
    return {"columnas": columnas, "filas": datos, "meta": meta}


def _meta_xlsx(ruta: str) -> dict:
    """Extrae solo {estado, candidato} de un XLSX de bloque (para el listado)."""
    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.active
        meta: dict = {}
        for fila in ws.iter_rows(values_only=True):
            if not fila or fila[0] is None:
                continue
            et = str(fila[0]).strip().lower()
            if et in _META_LABELS:
                meta[_META_LABELS[et]] = (
                    str(fila[1]) if len(fila) > 1 and fila[1] is not None else ""
                )
        wb.close()
        return meta
    except Exception:
        return {}


def leer_detalle(ruta: str) -> dict:
    """Devuelve el contenido parseado de un resultado (JSON manual o XLSX bloque)."""
    p = Path(ruta)
    if not p.exists():
        raise FileNotFoundError(f"No existe: {ruta}")
    if p.suffix.lower() == ".json":
        with open(p, encoding="utf-8") as fh:
            data = json.load(fh)
        return {"tipo": "manual", "data": data}
    if p.suffix.lower() == ".xlsx":
        return {"tipo": "block", "data": _leer_xlsx(ruta)}
    raise ValueError(f"Formato no soportado: {p.suffix}")
