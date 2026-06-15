# -*- coding: utf-8 -*-
"""
Comparación batch: Fuerza Bruta k-MIP (N ≤ 6) vs KQNodes vs KGeoMIP.

Procesa ÚNICAMENTE las muestras pequeñas de Brute_Force/samples_force/ (N ≤ 6),
que es donde la fuerza bruta exhaustiva es tratable. Los archivos con N > 6 se
omiten: la fuerza bruta no se compara contra samples_binary ni ningún N grande.

Por cada archivo se generan VARIAS pruebas (configuraciones distintas de
alcance/mecanismo, hasta `N_PRUEBAS_POR_ARCHIVO`), de modo que haya muchas
pruebas por cada k y por cada N — no una sola como antes. Parámetros fijos por
prueba:
  estado    = "0" * N         (todos ceros)
  candidato = "1" * N         (sistema completo de fondo)
  alcance / mecanismo         ← varían por prueba (subconjuntos no vacíos)

El muestreo de (alcance, mecanismo) es determinista (seed fijo) y siempre incluye
primero el sistema completo ('1'*N, '1'*N). Para N pequeños cuyo espacio total de
combinaciones es menor que `N_PRUEBAS_POR_ARCHIVO` (p. ej. N=2 → 9 combos) se usan
todas las posibles.

BruteForce y KQNodes corren en el proceso actual (KQNodes/src en sys.path).
KGeoMIP corre en subproceso aislado vía data/_worker_motor.py para evitar el
conflicto de nombres de paquete 'src' entre KGeoMIP y KQNodes. Para no pagar el
arranque del subproceso una vez por prueba, TODAS las pruebas de un mismo k se
mandan en UN solo subproceso (el worker acepta una lista de tests).

Salida: Brute_Force/results/comparacion_fuerza_bruta_<fecha>.xlsx
  Hoja "BF vs KQNodes vs KGeoMIP"  — comparación por (prueba, k) para N ≤ 6
  Hoja "KQNodes vs KGeoMIP"        — comparación QN vs GEO para esos mismos casos

Uso:
    .venv/Scripts/python Brute_Force/comparar_fuerza_bruta.py
"""

import contextlib
import json
import logging
import os
import random
import subprocess
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# ── Rutas base ────────────────────────────────────────────────────────────────
BRUTE_FORCE_ROOT = Path(__file__).resolve().parent   # AYDA_2026_1/Brute_Force
REPO_ROOT   = BRUTE_FORCE_ROOT.parent
QNODES_ROOT = REPO_ROOT / "KQNodes"
GEOMIP_ROOT = REPO_ROOT / "KGeoMIP"
DATA_ROOT   = REPO_ROOT / "data"
WORKER      = DATA_ROOT / "_worker_motor.py"
PYTHON      = sys.executable
RESULTS_DIR = BRUTE_FORCE_ROOT / "results"
SAMPLES_FRC = BRUTE_FORCE_ROOT / "samples_force"

RESULT_SENTINEL  = "@@RESULT@@"
TOLERANCIA       = 1e-9   # para considerar dos Phi "idénticos"
TOLERANCIA_COTA  = 1e-6   # margen para fallo de cota (evita falsos positivos por float32 vs float64)
N_MAX_BF         = 6      # límite práctico de la fuerza bruta

# Cuántas pruebas (configuraciones de alcance/mecanismo) generar por archivo.
# Si el espacio total de combinaciones no vacías es menor, se usan todas.
N_PRUEBAS_POR_ARCHIVO = 20
SEED_PRUEBAS          = 2026

# samples_force/ → SOLO N pequeños (N ≤ 6) — BF + KQNodes + KGeoMIP.
# La fuerza bruta NO se ejecuta sobre samples_binary ni N grandes (intratable).

# ── Importar KQNodes en proceso ───────────────────────────────────────────────
if str(QNODES_ROOT) not in sys.path:
    sys.path.insert(0, str(QNODES_ROOT))

from src.models.base.application import aplicacion       # noqa: E402
from src.middlewares.profile import gestor_perfilado     # noqa: E402
from src.strategies.force import BruteForceKMIP          # noqa: E402
from src.strategies.q_nodes import QNodes                # noqa: E402

aplicacion.desactivar_profiling()
gestor_perfilado.enabled = False


# ── Supresión de output de los motores ────────────────────────────────────────
@contextlib.contextmanager
def _silenciar():
    devnull = open(os.devnull, "w", encoding="utf-8")
    nivel_previo = logging.root.manager.disable
    logging.disable(logging.CRITICAL)
    try:
        with contextlib.redirect_stdout(devnull), contextlib.redirect_stderr(devnull):
            yield
    finally:
        devnull.close()
        logging.disable(nivel_previo)


# ── Utilidades ────────────────────────────────────────────────────────────────
def sep(titulo: str) -> None:
    print(f"\n{'═' * 68}")
    print(f"  {titulo}")
    print("═" * 68)


def cargar_tpm(csv_path: Path) -> tuple:
    with open(csv_path, "r", encoding="utf-8") as fh:
        tpm = np.genfromtxt(fh, delimiter=",")
    return tpm, int(tpm.shape[1])


def generar_pruebas(n: int, cantidad: int, seed: int = SEED_PRUEBAS) -> list:
    """Genera hasta `cantidad` pruebas (alcance, mecanismo) distintas para N nodos.

    Cada máscara es un string binario de longitud n con al menos un bit en 1
    (subconjunto no vacío de nodos). Siempre incluye primero el sistema completo
    ('1'*n, '1'*n). Si el espacio total de combinaciones es ≤ `cantidad`, devuelve
    todas. El muestreo es determinista (seed fijo + n) para reproducibilidad.
    """
    # El alcance necesita ≥ 2 nodos activos para que k=2 sea válido.
    # El mecanismo puede ser cualquier subconjunto no vacío.
    alc_validos = [format(i, f"0{n}b") for i in range(1, 2 ** n) if bin(i).count("1") >= 2]
    mec_validos = [format(i, f"0{n}b") for i in range(1, 2 ** n)]
    full = ("1" * n, "1" * n)
    combos = [(a, m) for a in alc_validos for m in mec_validos]
    if len(combos) <= cantidad:
        combos = [c for c in sorted(combos) if c != full]
        return [full] + combos
    rng = random.Random(seed + n)
    resto = [c for c in combos if c != full]
    muestra = sorted(rng.sample(resto, cantidad - 1))
    return [full] + muestra


def comparar_phi(phi_bf: float, phi_heur: float) -> tuple:
    """Devuelve (igual, cota_ok, err_rel) comparando heurística contra bruta.

    cota_ok usa TOLERANCIA_COTA (1e-6) para evitar falsos positivos por
    diferencias de precisión float32/float64 en la acumulación de la EMD.
    """
    igual    = abs(phi_bf - phi_heur) <= TOLERANCIA
    cota_ok  = phi_heur >= phi_bf - TOLERANCIA_COTA
    err_rel  = abs(phi_bf - phi_heur) / phi_bf if phi_bf > TOLERANCIA else 0.0
    return igual, cota_ok, round(err_rel, 9)


# ── Motores ───────────────────────────────────────────────────────────────────

def correr_bruta(tpm: np.ndarray, n: int, estado: str, candidato: str,
                 alcance: str, mecanismo: str, k: Optional[int],
                 permitir_vacio: bool) -> dict:
    """Fuerza bruta exhaustiva k-MIP (BruteForceKMIP, solo N ≤ 6)."""
    try:
        t0 = time.time()
        with _silenciar():
            bf  = BruteForceKMIP(tpm)
            sol = bf.aplicar_estrategia(
                estado_inicial=estado,
                condicion=candidato,
                alcance=alcance,
                mecanismo=mecanismo,
                k=k,
                permitir_presente_vacio=permitir_vacio,
                umbral_configuraciones=500_000,
                max_futuros=N_MAX_BF,
            )
        return {
            "ok": True,
            "perdida": float(sol.perdida),
            "particion": str(sol.particion),
            "tiempo": float(getattr(sol, "tiempo_total", time.time() - t0)),
            "optimos_por_k": {kk: round(float(v[0]), 9)
                               for kk, v in bf.optimos_por_k.items()},
        }
    except Exception as exc:
        return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}


def correr_qnodes(tpm: np.ndarray, n: int, estado: str, candidato: str,
                  alcance: str, mecanismo: str, k: Optional[int],
                  permitir_vacio: bool) -> dict:
    """KQNodes en proceso."""
    try:
        t0 = time.time()
        with _silenciar():
            q   = QNodes(tpm)
            sol = q.aplicar_estrategia(
                estado_inicial=estado,
                condicion=candidato,
                alcance=alcance,
                mecanismo=mecanismo,
                k=k,
                permitir_presente_vacio=permitir_vacio,
            )
        return {
            "ok": True,
            "perdida": float(sol.perdida),
            "particion": str(sol.particion),
            "tiempo": float(getattr(sol, "tiempo_total", time.time() - t0)),
        }
    except Exception as exc:
        return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}


def correr_geomip_lote(csv_path: Path, n: int, estado: str, candidato: str,
                       k: int, pruebas: list, permitir_vacio: bool) -> dict:
    """KGeoMIP en subproceso aislado para TODAS las pruebas de un mismo k.

    Manda la lista completa de tests al worker (un único arranque del motor) y
    devuelve {indice_prueba (1-based): payload}. payload trae ok/perdida/tiempo/
    particion o ok=False/error.
    """
    tests = [{"row": i, "alcance": a, "mecanismo": m}
             for i, (a, m) in enumerate(pruebas, 1)]
    cfg = {
        "engine": "geomip",
        "root": str(GEOMIP_ROOT),
        # Ruta ABSOLUTA: el worker hace os.chdir(root=KGeoMIP) antes de leer la
        # TPM, así que una ruta relativa al CWD original quedaría rota.
        "tpm": str(Path(csv_path).resolve()),
        "estado": estado,
        "candidato": candidato,
        "permitir_presente_vacio": permitir_vacio,
        "k": k,
        "tests": tests,
    }
    with tempfile.NamedTemporaryFile("w", suffix=".json",
                                     delete=False, encoding="utf-8") as tf:
        json.dump(cfg, tf, ensure_ascii=False)
        lote = tf.name

    resultados: dict = {}
    try:
        # Leemos toda la salida de una vez (communicate). El patrón de lectura
        # incremental por líneas con bufsize=1 resultó intermitente en Windows
        # (a veces el pipe se cierra antes de leer nada); communicate es estable.
        proc = subprocess.run(
            [PYTHON, str(WORKER), lote],
            stdout=subprocess.PIPE, stderr=subprocess.DEVNULL,
            text=True, encoding="utf-8", errors="replace",
        )
        for linea in proc.stdout.splitlines():
            if not linea.startswith(RESULT_SENTINEL):
                continue
            payload = json.loads(linea[len(RESULT_SENTINEL):])
            if "fatal" in payload:
                return {i: {"ok": False, "error": payload["fatal"]}
                        for i in range(1, len(pruebas) + 1)}
            if "ready" in payload:
                continue
            resultados[payload["row"]] = payload
    finally:
        try:
            os.unlink(lote)
        except OSError:
            pass

    # Pruebas sin respuesta → error explícito.
    for i in range(1, len(pruebas) + 1):
        resultados.setdefault(i, {"ok": False, "error": "sin respuesta del worker"})
    return resultados


# ── Exportación XLSX ──────────────────────────────────────────────────────────

def _escribir_cabeceras(ws, cabeceras: list) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hf   = Font(bold=True, color="FFFFFF")
    hfll = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, titulo in enumerate(cabeceras, 1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.font, c.fill, c.alignment = hf, hfll, ctr
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(titulo) + 2)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def exportar_xlsx(filas_bf: list, filas_comp: list) -> Path:
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment

    verde   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    naranja = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    rojo    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ctr     = Alignment(horizontal="center", vertical="center")

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    fecha = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    ruta  = RESULTS_DIR / f"comparacion_fuerza_bruta_{fecha}.xlsx"
    wb    = openpyxl.Workbook()

    # ── Hoja 1: BF vs KQNodes vs KGeoMIP ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "BF vs KQNodes vs KGeoMIP"
    hdrs1 = [
        "Archivo", "N", "Prueba", "Alcance", "Mecanismo", "k",
        "Φ_FuerzaBruta", "Φ_KQNodes", "Φ_KGeoMIP",
        "QN_≥_BF", "GEO_≥_BF",
        "err_rel_QN", "err_rel_GEO",
        "Φ_igual_QN", "Φ_igual_GEO",
        "t_QNodes(s)", "t_KGeoMIP(s)",
        "Nota_QN", "Nota_GEO",
    ]
    _escribir_cabeceras(ws1, hdrs1)

    for ri, f in enumerate(filas_bf, 2):
        row = [
            f.get("archivo"), f.get("n"), f.get("prueba"),
            f.get("alcance"), f.get("mecanismo"), f.get("k"),
            f.get("phi_bf"), f.get("phi_qn"), f.get("phi_geo"),
            f.get("qn_cota_ok"), f.get("geo_cota_ok"),
            f.get("err_rel_qn"), f.get("err_rel_geo"),
            f.get("phi_igual_qn"), f.get("phi_igual_geo"),
            f.get("t_qn"), f.get("t_geo"),
            f.get("nota_qn", ""), f.get("nota_geo", ""),
        ]
        if f.get("qn_cota_ok") is False or f.get("geo_cota_ok") is False:
            fill = rojo
        elif f.get("phi_igual_qn") and f.get("phi_igual_geo"):
            fill = verde
        else:
            fill = naranja
        for ci, val in enumerate(row, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.alignment = ctr
            c.fill = fill

    # ── Hoja 2: KQNodes vs KGeoMIP ───────────────────────────────────────────
    ws2 = wb.create_sheet("KQNodes vs KGeoMIP")
    hdrs2 = [
        "Archivo", "N", "Prueba", "Alcance", "Mecanismo", "k",
        "Φ_KQNodes", "Φ_KGeoMIP",
        "Coinciden", "diff_abs", "err_rel",
        "t_QNodes(s)", "t_KGeoMIP(s)",
        "Nota_QN", "Nota_GEO",
    ]
    _escribir_cabeceras(ws2, hdrs2)

    for ri, f in enumerate(filas_comp, 2):
        row = [
            f.get("archivo"), f.get("n"), f.get("prueba"),
            f.get("alcance"), f.get("mecanismo"), f.get("k", "libre"),
            f.get("phi_qn"), f.get("phi_geo"),
            f.get("coinciden"),
            f.get("diff_abs"), f.get("err_rel"),
            f.get("t_qn"), f.get("t_geo"),
            f.get("nota_qn", ""), f.get("nota_geo", ""),
        ]
        if f.get("coinciden"):
            fill = verde
        elif (f.get("err_rel") or 0) < 0.01:
            fill = naranja
        else:
            fill = rojo
        for ci, val in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.alignment = ctr
            c.fill = fill

    wb.save(ruta)
    return ruta


# ── Punto de entrada ──────────────────────────────────────────────────────────

def main() -> None:
    sep("Comparación: Fuerza Bruta vs KQNodes vs KGeoMIP")

    # samples_force/ = N pequeños (N≤6) → comparación completa con fuerza bruta.
    archivos_frc = sorted(SAMPLES_FRC.glob("N*.csv"))

    if not archivos_frc:
        print(f" No se encontraron CSV en {SAMPLES_FRC}.")
        sys.exit(1)

    print(f"\n  samples_force  : {len(archivos_frc)} archivo(s)  [BF + KQNodes + KGeoMIP, N ≤ {N_MAX_BF}]")
    for f in archivos_frc:
        print(f"    {f.name}")

    opcion = input(
        "\n ¿Permitir mecanismo vacío (∅) en las partes?\n"
        "   1. Sí (recomendado) — lower bound real para la bruta\n"
        "   2. No\n"
        " Seleccione (1/2) o [ENTER] para Sí: "
    ).strip()
    permitir_vacio = (opcion != "2")
    print(f" permitir_presente_vacio = {permitir_vacio}")

    txt_n = input(
        f"\n ¿Cuántas pruebas (alcance/mecanismo) por archivo? [{N_PRUEBAS_POR_ARCHIVO}]: "
    ).strip()
    try:
        n_pruebas = max(1, int(txt_n)) if txt_n else N_PRUEBAS_POR_ARCHIVO
    except ValueError:
        n_pruebas = N_PRUEBAS_POR_ARCHIVO
    print(f" pruebas por archivo = {n_pruebas} (las posibles si el espacio es menor)")

    filas_bf: list   = []
    filas_comp: list = []

    # ── samples_force: BF + QNodes + KGeoMIP por (prueba, k) (N ≤ 6) ─────────
    sep("samples_force — BruteForce + KQNodes + KGeoMIP (N ≤ 6)")

    for csv_path in archivos_frc:
        tpm, n = cargar_tpm(csv_path)
        nombre = csv_path.name
        estado = "0" * n
        candidato = "1" * n

        if n > N_MAX_BF:
            print(f"\n  [{nombre}]  N={n} > {N_MAX_BF} — OMITIDO (la fuerza bruta solo compara N ≤ {N_MAX_BF}).")
            continue

        pruebas = generar_pruebas(n, n_pruebas)
        k_values = list(range(2, n + 1))
        print(f"\n  [{nombre}]  N={n}  ·  {len(pruebas)} pruebas  ·  k ∈ {k_values}")

        # 1) BF + QNodes (ambos in-process) por cada prueba.
        bf_por_prueba: dict  = {}
        qn_por_prueba: dict  = {}  # (i, k) → resultado QNodes
        for i, (alcance, mecanismo) in enumerate(pruebas, 1):
            print(f"    prueba {i:>2}/{len(pruebas)}  alc={alcance} mec={mecanismo}  BF...",
                  end=" ", flush=True)
            bf_res = correr_bruta(tpm, n, estado, candidato, alcance, mecanismo,
                                  None, permitir_vacio)
            bf_por_prueba[i] = bf_res
            print("OK" if bf_res.get("ok") else f"ERR({bf_res.get('error','?')})", end="  ", flush=True)
            print("QN...", end=" ", flush=True)
            for k in k_values:
                # alcance.count("1") = n_dims del subsistema; k > n_dims es inválido.
                if k > alcance.count("1"):
                    qn_por_prueba[(i, k)] = {"ok": False, "error": f"k={k} > nodos alcance ({alcance.count('1')})"}
                    continue
                qn_por_prueba[(i, k)] = correr_qnodes(
                    tpm, n, estado, candidato, alcance, mecanismo, k, permitir_vacio)
            print("OK")

        # 2) KGeoMIP en subproceso, UN arranque por k (solo pruebas con alcance ≥ k bits).
        geo_por_k: dict = {}
        for k in k_values:
            pruebas_k = [(a, m) for a, m in pruebas if a.count("1") >= k]
            if not pruebas_k:
                geo_por_k[k] = {}
                continue
            print(f"    KGeoMIP lote k={k} ({len(pruebas_k)} pruebas válidas)...", end=" ", flush=True)
            # El índice de row que devuelve el worker es 1-based sobre pruebas_k,
            # así que mapeamos de vuelta al índice original de pruebas.
            idx_map = {j + 1: pruebas.index(p) + 1 for j, p in enumerate(pruebas_k)}
            raw = correr_geomip_lote(csv_path, n, estado, candidato, k, pruebas_k, permitir_vacio)
            geo_por_k[k] = {idx_map[j]: v for j, v in raw.items()}
            print("OK")

        # 3) Ensamblar filas por (prueba, k); omitir si alcance < k nodos.
        for i, (alcance, mecanismo) in enumerate(pruebas, 1):
            bf_res = bf_por_prueba[i]
            optimos = bf_res.get("optimos_por_k", {}) if bf_res.get("ok") else {}
            for k in k_values:
                if k > alcance.count("1"):
                    continue  # subsistema demasiado pequeño para esta k
                phi_bf = optimos.get(k)
                qn_res = qn_por_prueba[(i, k)]
                geo_res = geo_por_k[k].get(i, {"ok": False, "error": "sin dato"})

                phi_qn  = float(qn_res["perdida"])  if qn_res.get("ok")  else None
                phi_geo = float(geo_res["perdida"]) if geo_res.get("ok") else None
                t_qn    = round(float(qn_res.get("tiempo", 0)), 4)  if phi_qn  is not None else None
                t_geo   = round(float(geo_res.get("tiempo", 0)), 4) if phi_geo is not None else None

                qn_igual = qn_cota_ok = err_qn = None
                geo_igual = geo_cota_ok = err_geo = None
                if phi_bf is not None and phi_qn is not None:
                    qn_igual, qn_cota_ok, err_qn = comparar_phi(phi_bf, phi_qn)
                if phi_bf is not None and phi_geo is not None:
                    geo_igual, geo_cota_ok, err_geo = comparar_phi(phi_bf, phi_geo)

                filas_bf.append({
                    "archivo": nombre, "n": n, "prueba": i,
                    "alcance": alcance, "mecanismo": mecanismo, "k": k,
                    "phi_bf":  round(phi_bf, 9) if phi_bf is not None else "ERROR",
                    "phi_qn":  round(phi_qn, 9)  if phi_qn  is not None else "ERROR",
                    "phi_geo": round(phi_geo, 9) if phi_geo is not None else "ERROR",
                    "qn_cota_ok": qn_cota_ok, "geo_cota_ok": geo_cota_ok,
                    "err_rel_qn": err_qn, "err_rel_geo": err_geo,
                    "phi_igual_qn": qn_igual, "phi_igual_geo": geo_igual,
                    "t_qn": t_qn, "t_geo": t_geo,
                    "nota_qn":  "" if qn_res.get("ok")  else qn_res.get("error", ""),
                    "nota_geo": "" if geo_res.get("ok") else geo_res.get("error", ""),
                })

                # Comparativa QN vs GEO (independiente de la bruta).
                if phi_qn is not None and phi_geo is not None:
                    diff_abs = abs(phi_qn - phi_geo)
                    mn = min(phi_qn, phi_geo)
                    filas_comp.append({
                        "archivo": nombre, "n": n, "prueba": i,
                        "alcance": alcance, "mecanismo": mecanismo, "k": k,
                        "phi_qn": round(phi_qn, 9), "phi_geo": round(phi_geo, 9),
                        "coinciden": diff_abs <= TOLERANCIA,
                        "diff_abs": round(diff_abs, 9),
                        "err_rel":  round(diff_abs / mn, 9) if mn > TOLERANCIA else 0.0,
                        "t_qn": t_qn, "t_geo": t_geo,
                        "nota_qn": "", "nota_geo": "",
                    })

    # ── Resumen por consola ───────────────────────────────────────────────────
    sep("Resumen")
    if filas_bf:
        total = len(filas_bf)
        iguales_qn  = sum(1 for f in filas_bf if f.get("phi_igual_qn"))
        iguales_geo = sum(1 for f in filas_bf if f.get("phi_igual_geo"))
        fallos_qn   = sum(1 for f in filas_bf if f.get("qn_cota_ok") is False)
        fallos_geo  = sum(1 for f in filas_bf if f.get("geo_cota_ok") is False)
        print(f"  BF vs Frameworks — {total} combinaciones (archivo × prueba × k):")
        print(f"    KQNodes  — Φ exacto: {iguales_qn}/{total}   Fallos cota: {fallos_qn}")
        print(f"    KGeoMIP  — Φ exacto: {iguales_geo}/{total}   Fallos cota: {fallos_geo}")
        if fallos_qn or fallos_geo:
            print("    ⚠  Hay fallos de cota inferior — revisar implementación.")
    else:
        print("  No se generaron filas de comparación BF vs Frameworks.")

    if filas_comp:
        total_c = len(filas_comp)
        coinciden = sum(1 for f in filas_comp if f.get("coinciden"))
        print(f"\n  KQNodes vs KGeoMIP — {total_c} combinaciones:")
        print(f"    Coinciden exactamente: {coinciden}/{total_c}")

    # ── Exportar XLSX ─────────────────────────────────────────────────────────
    sep("Exportando XLSX")
    try:
        ruta = exportar_xlsx(filas_bf, filas_comp)
        print(f"  ✓ Guardado en: {ruta}")
    except Exception as exc:
        print(f"  ✗ Error al guardar XLSX: {exc}")


if __name__ == "__main__":
    main()
