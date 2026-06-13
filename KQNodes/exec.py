# -*- coding: utf-8 -*-
"""
Script de ejecución interactivo para KQNodes — k-Partición de Mínima Información (DP + B&B)

Modos de ejecución:
  1. Manual   — ingreso interactivo de una sola prueba.
  2. Por bloque — carga un CSV con múltiples pruebas y guarda los resultados
                  en el destino que el usuario indique.

Formato del CSV para modo bloque:
  #Prueba,Alcance o Purview (t+1),Mecanismo(t)
  1,ABCDE,ABCDE
  2,ABCDE,ABCD

  - Alcance / Mecanismo: etiquetas de nodos (A=nodo 0, B=nodo 1, ...).
  - El candidato y el estado inicial se ingresan una sola vez para todo el lote.
  - El resultado se guarda automáticamente en KQNodes/results/block/ como .xlsx.
"""

import sys
import csv
import json
import random
import re
import time
from pathlib import Path
from datetime import datetime

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.models.base.application import aplicacion
from src.strategies.q_nodes import DynamicPartition, PERMITIR_PRESENTE_VACIO_POR_DEFECTO
from src.funcs.iit import ABECEDARY
from src.middlewares.profile import gestor_perfilado

QNODES_ROOT           = Path(__file__).resolve().parent
# Los datasets/TPMs y CSV de pruebas viven en la carpeta data/ de la RAÍZ del
# repositorio (compartida con KGeoMIP); los resultados sí van a KQNodes/results.
REPO_ROOT             = QNODES_ROOT.parent
DATA_ROOT             = REPO_ROOT / "data"
SAMPLES_BINARY_DIR    = DATA_ROOT / "samples_binary"
SAMPLES_NO_BINARY_DIR = DATA_ROOT / "samples_no_binary"
PRUEBAS_DIR           = DATA_ROOT / "Pruebas"


# ── Utilidades de presentación ─────────────────────────────────────────────

def sep(titulo: str):
    print(f"\n{'='*60}")
    print(f"  {titulo}")
    print('='*60)


def formatear_tiempo(segundos: float) -> str:
    """Convierte segundos en cadena legible: s / min s / h min s."""
    if segundos < 60:
        return f"{segundos:.4f} s"
    elif segundos < 3600:
        mins = int(segundos // 60)
        segs = segundos % 60
        return f"{mins} min {segs:.2f} s"
    else:
        horas = int(segundos // 3600)
        resto = segundos % 3600
        mins = int(resto // 60)
        segs = resto % 60
        return f"{horas} h {mins} min {segs:.2f} s"


# ── Helpers de entrada ─────────────────────────────────────────────────────

def _seleccionar_tpm() -> tuple:
    """Pide N y tipo (binaria/no binaria) en terminal, lista los archivos que coincidan y deja elegir.
    Opción 0 para volver a ingresar N. Retorna (ruta, tpm, n_nodos)."""
    while True:
        n_str = input("\n Ingrese el tamaño N (número de nodos) de la TPM a cargar: ").strip()
        if not n_str.isdigit():
            print(" Error: ingrese un número entero válido.")
            continue
        n = int(n_str)

        while True:
            tipo = input(
                "\n ¿Qué tipo de TPM desea cargar?\n"
                "   1. Binaria      — carpeta samples_binary\n"
                "   2. No binaria   — carpeta samples_no_binary\n"
                " Seleccione (1 o 2): "
            ).strip()
            if tipo in ("1", "2"):
                break
            print(" Error: ingrese 1 o 2.")

        directorio_tpm = SAMPLES_BINARY_DIR if tipo == "1" else SAMPLES_NO_BINARY_DIR

        patron   = re.compile(rf'^N{n}\D', re.IGNORECASE)
        archivos = sorted(f for f in directorio_tpm.glob("*.csv") if patron.match(f.name))

        if not archivos:
            disponibles = sorted(directorio_tpm.glob("*.csv"))
            print(f"\n No se encontraron archivos con N={n} en '{directorio_tpm.name}'.")
            if disponibles:
                print(f" Archivos disponibles: {[f.name for f in disponibles]}")
            print(" Intente con otro N.")
            continue

        print(f"\n Archivos disponibles para N={n} ({directorio_tpm.name}):")
        for i, archivo in enumerate(archivos, 1):
            print(f"   {i}. {archivo.name}")
        print(f"   0. Volver a elegir N")

        while True:
            sel = input(f" Seleccione un archivo [1-{len(archivos)}] o 0 para volver: ").strip()
            if sel == "0":
                break
            if sel.isdigit() and 1 <= int(sel) <= len(archivos):
                ruta = archivos[int(sel) - 1]
                try:
                    tpm     = np.genfromtxt(ruta, delimiter=',')
                    n_nodos = tpm.shape[1]
                except Exception as e:
                    print(f" Error al leer la TPM: {e}")
                    sys.exit(1)
                return ruta, tpm, n_nodos
            print(f" Error: seleccione un número entre 1 y {len(archivos)}, o 0 para volver.")


def _seleccionar_csv_pruebas(n_nodos: int) -> Path:
    """Lista los archivos de Pruebas que coincidan con N y deja elegir.
    Opción 0 para volver a ingresar N."""
    n = n_nodos
    while True:
        patron   = re.compile(rf'[_\-]N{n}[_\-.]', re.IGNORECASE)
        archivos = sorted(f for f in PRUEBAS_DIR.glob("*.csv") if patron.search(f.name))

        if not archivos:
            disponibles = sorted(PRUEBAS_DIR.glob("*.csv"))
            print(f"\n No se encontraron archivos de pruebas para N={n} en '{PRUEBAS_DIR.name}'.")
            if disponibles:
                print(f" Archivos disponibles: {[f.name for f in disponibles]}")
        else:
            print(f"\n Archivos de pruebas disponibles para N={n}:")
            for i, archivo in enumerate(archivos, 1):
                print(f"   {i}. {archivo.name}")
            print(f"   0. Volver a elegir N")

            while True:
                sel = input(f" Seleccione un archivo [1-{len(archivos)}] o 0 para volver: ").strip()
                if sel == "0":
                    break
                if sel.isdigit() and 1 <= int(sel) <= len(archivos):
                    return archivos[int(sel) - 1]
                print(f" Error: seleccione un número entre 1 y {len(archivos)}, o 0 para volver.")

        n_nuevo = input(" Ingrese el N deseado para las pruebas (o [ENTER] para cancelar): ").strip()
        if not n_nuevo:
            print(" Ejecución cancelada.")
            sys.exit(1)
        if n_nuevo.isdigit():
            n = int(n_nuevo)
        else:
            print(" Error: ingrese un número entero válido.")


def _pedir_candidato(n_nodos: int) -> tuple:
    """Solicita el sistema candidato/condición. Retorna (bits_candidato, n_candidatos)."""
    print(f"\n La red consta de {n_nodos} variables lógicas asociadas.")
    candidato_input = input(
        f" Ingrese el SISTEMA CANDIDATO en binario (longitud: {n_nodos} bits)\n"
        f"   O presione [ENTER] para tomar TODO el sistema: "
    ).strip()

    if not candidato_input:
        bits_candidato = "1" * n_nodos
        print(f" Seleccionado el sistema completo: {bits_candidato}")
    else:
        if set(candidato_input).issubset({"0", "1"}) and len(candidato_input) == n_nodos:
            bits_candidato = candidato_input
        else:
            print(f" Error: La máscara debe ser cadena de unos y ceros de longitud {n_nodos}.")
            sys.exit(1)

    n_candidatos = bits_candidato.count("1")
    if n_candidatos < 2:
        print(" Error: El sistema candidato debe tener al menos 2 variables (1s).")
        sys.exit(1)

    return bits_candidato, n_candidatos


def _pedir_estado(n_nodos: int) -> str:
    """Solicita el estado inicial del sistema completo (n_nodos bits)."""
    estado_input = input(
        f"\n Ingrese el ESTADO INICIAL en binario (longitud: {n_nodos} bits)\n"
        f"   O presione [ENTER] para uno RANDOM: "
    ).strip()

    if not estado_input:
        estado = "".join(str(random.randint(0, 1)) for _ in range(n_nodos))
        print(f" Estado aleatorio auto-generado: {estado}")
    else:
        if set(estado_input).issubset({"0", "1"}) and len(estado_input) == n_nodos:
            estado = estado_input
        else:
            print(f" Error: El estado debe ser una cadena de {n_nodos} bits (0s y 1s).")
            sys.exit(1)

    return estado


def _pedir_k(n_candidatos: int) -> "int | None":
    """Solicita el valor de K. Retorna entero o None (= buscar todas las k)."""
    while True:
        print(f"\n Número máximo de particiones posibles: {n_candidatos}")
        k_input = input(
            f" Ingrese el valor de K (particiones a evaluar) [2-{n_candidatos}]\n"
            f"   O presione [ENTER] para evaluar TODAS las k posibles: "
        ).strip()

        if not k_input:
            print(f" Se evaluarán TODAS las k posibles (2 a {n_candidatos}).")
            return None

        if not k_input.isdigit():
            print(" Error: K debe ser un número entero. Intenta de nuevo.")
            continue

        k_val = int(k_input)
        if k_val < 2 or k_val > n_candidatos:
            print(f" Error: K debe estar entre 2 y {n_candidatos}. Intenta de nuevo.")
            continue

        print(f" Se evaluará K = {k_val}.")
        return k_val


def _pedir_permitir_presente_vacio() -> bool:
    """Pregunta si se permite mecanismo vacío (∅) en las partes."""
    opcion = input(
        "\n ¿Permitir mecanismo vacío (∅) en las partes?\n"
        "   1. No (por defecto) — cada parte usa el mecanismo completo de sus nodos\n"
        "   2. Sí — permite que algunas partes usen mecanismo ∅ (mayor exploración)\n"
        " Seleccione (1 o 2) o [ENTER] para No: "
    ).strip()
    if opcion == "2":
        print(" Mecanismo vacío (∅) HABILITADO.")
        return True
    print(" Mecanismo vacío (∅) deshabilitado (por defecto).")
    return False


def _letras_a_binario(etiquetas: str, n_nodos: int) -> str:
    """
    Convierte etiquetas de nodos (ej: 'ABCEFG') a máscara binaria de longitud n_nodos.
    Parsea con matching greedy para soportar etiquetas multi-caracter (N>26).
    """
    etiquetas = etiquetas.strip().upper()
    activos: set[str] = set()
    labels_validas = ABECEDARY[:n_nodos]
    idx = 0
    while idx < len(etiquetas):
        matched = False
        for length in range(min(3, len(etiquetas) - idx), 0, -1):
            lbl = etiquetas[idx:idx + length]
            if lbl in labels_validas:
                activos.add(lbl)
                idx += length
                matched = True
                break
        if not matched:
            idx += 1
    return "".join("1" if ABECEDARY[i] in activos else "0" for i in range(n_nodos))


# ── Guardar resultados — modo manual ──────────────────────────────────────

def _guardar_resultado_single(
    ruta_archivo: Path,
    estado: str,
    bits_candidato: str,
    bits_alcance: str,
    bits_mecanismo: str,
    solucion,
    k_elegido: "int | None",
    permitir_presente_vacio: bool,
):
    """Guarda el resultado de una ejecución manual en KQNodes/results/manual/."""
    results_dir = QNODES_ROOT / "results" / "manual"
    results_dir.mkdir(parents=True, exist_ok=True)

    fecha_actual = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    k_str = str(k_elegido) if k_elegido is not None else "All"
    nombre_archivo = f"resultado_{ruta_archivo.stem}_k{k_str}_{fecha_actual}.json"
    ruta_salida = results_dir / nombre_archivo

    res_data = {
        "dataset": ruta_archivo.name,
        "estado_inicial": estado,
        "sistema_candidato": bits_candidato,
        "alcance": bits_alcance,
        "mecanismo": bits_mecanismo,
        "k_solicitado": k_elegido,
        "permitir_presente_vacio": permitir_presente_vacio,
        "estrategia": solucion.estrategia,
        "perdida_phi": float(solucion.perdida),
        "distribucion_subsistema": (
            solucion.distribucion_subsistema.tolist()
            if hasattr(solucion.distribucion_subsistema, 'tolist')
            else solucion.distribucion_subsistema
        ),
        "distribucion_particion": (
            solucion.distribucion_particion.tolist()
            if hasattr(solucion.distribucion_particion, 'tolist')
            else solucion.distribucion_particion
        ),
        "particion": str(solucion.particion),
        "tiempo_busqueda_segundos": float(solucion.tiempo_ejecucion),
        "tiempo_preparacion_segundos": float(getattr(solucion, "tiempo_preparacion", 0.0)),
        "tiempo_total_segundos": float(solucion.tiempo_ejecucion),
        "tiempo_formateado": formatear_tiempo(float(solucion.tiempo_ejecucion)),
    }

    with open(ruta_salida, 'w', encoding='utf-8') as f:
        json.dump(res_data, f, indent=4, ensure_ascii=False)

    print(f"\n ✓ Resultados guardados en: {ruta_salida.relative_to(QNODES_ROOT)}")


# ── Escritura incremental de Excel (modo bloque) ──────────────────────────

COLUMNAS_BLOQUE = [
    ("#Prueba",                  10),
    ("Alcance o Purview (t+1)",  28),
    ("Mecanismo(t)",             22),
    ("Particion",                48),
    ("Perdida",                  16),
    ("Tiempo",                   20),
]

_HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
_HDR_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center")
_ALT_FILL  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_TOT_FILL  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def _crear_excel_bloque_inicial(ruta_salida: Path, pruebas_preparadas: list, n_pruebas: int) -> None:
    """
    Crea el Excel con todas las filas pre-rellenadas (solo #Prueba, Alcance, Mecanismo)
    y una fila final de Tiempo Total en pending. Los resultados se rellenan prueba a prueba.
    pruebas_preparadas debe estar en orden original del CSV (sorted by _orig_idx).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    for col_idx, (titulo, ancho) in enumerate(COLUMNAS_BLOQUE, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    center_align = Alignment(horizontal="center", vertical="center")
    pending_font = Font(italic=True, color="AAAAAA")

    for fila_idx, r in enumerate(pruebas_preparadas, start=2):
        relleno = _ALT_FILL if (fila_idx % 2 == 0) else None
        for col, val in [
            (1, r.get("#prueba")),
            (2, r.get("alcance_orig", "")),
            (3, r.get("mecanismo_orig", "")),
            (4, "pendiente..."),
            (5, None),
            (6, None),
        ]:
            c = ws.cell(row=fila_idx, column=col, value=val)
            c.alignment = center_align
            if relleno:
                c.fill = relleno
            if col == 4:
                c.font = pending_font
        ws.row_dimensions[fila_idx].height = 45

    # Fila de tiempo total al final
    fila_total = n_pruebas + 2
    bold_font = Font(bold=True)
    for col in range(1, 7):
        c = ws.cell(row=fila_total, column=col)
        c.fill = _TOT_FILL
        c.alignment = center_align
    ws.cell(row=fila_total, column=1, value="Tiempo Total Lote").font = bold_font
    ws.cell(row=fila_total, column=1).fill = _TOT_FILL
    ws.cell(row=fila_total, column=1).alignment = center_align
    ws.cell(row=fila_total, column=6, value="calculando...").font = Font(italic=True, color="AAAAAA")
    ws.cell(row=fila_total, column=6).fill = _TOT_FILL
    ws.cell(row=fila_total, column=6).alignment = center_align
    ws.row_dimensions[fila_total].height = 22

    # Fila de "arranque del motor" (warmup): tiempo de preparar el subsistema /
    # tabla — el coste de "calentar motores", aparte del tiempo de las pruebas.
    fila_warm = fila_total + 1
    warm_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    for col in range(1, 7):
        c = ws.cell(row=fila_warm, column=col)
        c.fill = warm_fill
        c.alignment = center_align
    ws.cell(row=fila_warm, column=1, value="Arranque motor (warmup)").font = bold_font
    ws.cell(row=fila_warm, column=1).fill = warm_fill
    ws.cell(row=fila_warm, column=1).alignment = center_align
    ws.cell(row=fila_warm, column=6, value="calculando...").font = Font(italic=True, color="AAAAAA")
    ws.cell(row=fila_warm, column=6).fill = warm_fill
    ws.cell(row=fila_warm, column=6).alignment = center_align
    ws.row_dimensions[fila_warm].height = 22

    wb.save(ruta_salida)


def _actualizar_warmup_excel(ruta_salida: Path, fila_warm: int, tiempo_seg: float) -> None:
    """Actualiza la celda de tiempo de arranque del motor (warmup)."""
    try:
        wb = openpyxl.load_workbook(ruta_salida)
        ws = wb.active
        warm_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        c = ws.cell(row=fila_warm, column=6, value=formatear_tiempo(tiempo_seg))
        c.font = Font(bold=True)
        c.fill = warm_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        wb.save(ruta_salida)
    except Exception as e:
        print(f"   (Aviso: no se pudo actualizar el tiempo de arranque — {e})")


def _actualizar_fila_excel(ruta_salida: Path, excel_row: int, resultado: dict) -> None:
    """Abre el Excel, rellena las columnas de resultado de la fila indicada y lo guarda."""
    try:
        wb = openpyxl.load_workbook(ruta_salida)
        ws = wb.active

        part_align   = Alignment(wrap_text=True, vertical="top", horizontal="left")
        center_align = Alignment(horizontal="center", vertical="center")
        error_font   = Font(color="CC0000")

        if resultado.get("error"):
            c = ws.cell(row=excel_row, column=4, value=f"ERROR: {resultado['error']}")
            c.alignment = part_align
            c.font = error_font
            ws.cell(row=excel_row, column=5, value=None).alignment = center_align
            ws.cell(row=excel_row, column=6, value=None).alignment = center_align
        else:
            c = ws.cell(row=excel_row, column=4, value=resultado.get("particion", ""))
            c.alignment = part_align
            c.font = Font()
            perdida = resultado.get("perdida_phi")
            ws.cell(row=excel_row, column=5,
                    value=round(perdida, 6) if perdida is not None else None).alignment = center_align
            ws.cell(row=excel_row, column=6,
                    value=resultado.get("tiempo_formateado", "")).alignment = center_align

        wb.save(ruta_salida)
    except Exception as e:
        print(f"   (Aviso: no se pudo actualizar el Excel — {e})")


def _actualizar_tiempo_total_excel(ruta_salida: Path, fila_total: int, tiempo_seg: float) -> None:
    """Actualiza la celda de Tiempo Total en la fila de resumen del Excel."""
    try:
        wb = openpyxl.load_workbook(ruta_salida)
        ws = wb.active
        c = ws.cell(row=fila_total, column=6, value=formatear_tiempo(tiempo_seg))
        c.font = Font(bold=True)
        c.fill = _TOT_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        wb.save(ruta_salida)
    except Exception as e:
        print(f"   (Aviso: no se pudo actualizar el tiempo total — {e})")


# ── Modo 1: Ejecución manual ───────────────────────────────────────────────

def modo_manual(ruta_archivo: Path, tpm: np.ndarray, n_nodos: int):
    """Ejecuta una sola prueba de forma interactiva."""
    bits_candidato, n_candidatos = _pedir_candidato(n_nodos)
    estado = _pedir_estado(n_nodos)

    # ── Alcance ────────────────────────────────────────────────────────────
    print(f"\n Ingrese el ALCANCE (Purview t+1) en binario (longitud: {n_nodos} bits)")
    print(f"   Referencia — Candidato activo: {bits_candidato}")
    print(f"   1 = observar ese nodo en t+1  |  0 = ignorarlo (marginalizarlo)")
    alcance_input = input(
        f"   O presione [ENTER] para usar TODO el candidato como alcance: "
    ).strip()

    if not alcance_input:
        bits_alcance = bits_candidato
        print(f" Alcance: {bits_alcance} (igual al candidato)")
    else:
        if set(alcance_input).issubset({"0", "1"}) and len(alcance_input) == n_nodos:
            invalidos = [
                i for i, (a, c) in enumerate(zip(alcance_input, bits_candidato))
                if a == "1" and c == "0"
            ]
            if invalidos:
                print(f" Error: El alcance activa nodos {invalidos} que NO están en el candidato.")
                sys.exit(1)
            bits_alcance = alcance_input
        else:
            print(f" Error: Debe ser una cadena de 0s y 1s de longitud {n_nodos}.")
            sys.exit(1)

    # ── Mecanismo ──────────────────────────────────────────────────────────
    print(f"\n Ingrese el MECANISMO (t) en binario (longitud: {n_nodos} bits)")
    print(f"   Referencia — Candidato activo: {bits_candidato}")
    print(f"   1 = fijar ese nodo en su estado inicial  |  0 = marginalizarlo")
    mecanismo_input = input(
        f"   O presione [ENTER] para usar TODO el candidato como mecanismo: "
    ).strip()

    if not mecanismo_input:
        bits_mecanismo = bits_candidato
        print(f" Mecanismo: {bits_mecanismo} (igual al candidato)")
    else:
        if set(mecanismo_input).issubset({"0", "1"}) and len(mecanismo_input) == n_nodos:
            invalidos = [
                i for i, (m, c) in enumerate(zip(mecanismo_input, bits_candidato))
                if m == "1" and c == "0"
            ]
            if invalidos:
                print(f" Error: El mecanismo activa nodos {invalidos} que NO están en el candidato.")
                sys.exit(1)
            bits_mecanismo = mecanismo_input
        else:
            print(f" Error: Debe ser una cadena de 0s y 1s de longitud {n_nodos}.")
            sys.exit(1)

    # ── K y opciones ───────────────────────────────────────────────────────
    k_a_evaluar = _pedir_k(n_candidatos)
    permitir_vacio = _pedir_permitir_presente_vacio()

    # ── Resumen y ejecución ────────────────────────────────────────────────
    sep(" Iniciando Algoritmo — KQNodes k-Partición Óptima Global")
    print(f"   Dataset Cargado : {ruta_archivo.name}")
    print(f"   Sist. Candidato : {bits_candidato} ({n_candidatos} nodos)")
    print(f"   Estado Inicial  : {estado}")
    print(f"   Alcance (t+1)   : {bits_alcance}")
    print(f"   Mecanismo (t)   : {bits_mecanismo}")
    if k_a_evaluar:
        print(f"   K a evaluar     : {k_a_evaluar}")
    else:
        print(f"   K a evaluar     : Todas (2..{n_candidatos})")
    print(f"   Presente vacío  : {'Sí' if permitir_vacio else 'No'}\n")

    analizador = DynamicPartition(tpm)
    solucion = analizador.aplicar_estrategia(
        estado_inicial=estado,
        condicion=bits_candidato,
        alcance=bits_alcance,
        mecanismo=bits_mecanismo,
        k=k_a_evaluar,
        permitir_presente_vacio=permitir_vacio,
    )

    print(solucion)
    _guardar_resultado_single(
        ruta_archivo, estado, bits_candidato, bits_alcance, bits_mecanismo,
        solucion, k_a_evaluar, permitir_vacio,
    )


# ── Modo 2: Ejecución por bloque ───────────────────────────────────────────

def modo_bloque(ruta_tpm: Path, tpm: np.ndarray, n_nodos: int):
    """Ejecuta un lote de pruebas desde un CSV y guarda resultados en Excel."""
    sep(" Modo Bloque — Ejecución por Lotes desde CSV")

    print(
        "\n Formato esperado del CSV de pruebas:\n"
        "   #Prueba,Alcance o Purview (t+1),Mecanismo(t)\n"
        "   1,ABCDE,ABCDE\n"
        "   2,ABCDE,ABCD\n"
        "\n   - Alcance / Mecanismo: etiquetas de nodos (A=0, B=1, C=2, ...).\n"
    )

    bits_candidato, n_candidatos = _pedir_candidato(n_nodos)
    estado = _pedir_estado(n_nodos)

    # K global para todo el lote
    k_val_global = _pedir_k(n_candidatos)
    k_nombre = str(k_val_global) if k_val_global is not None else "All"

    # Opción de mecanismo vacío
    permitir_vacio = _pedir_permitir_presente_vacio()

    # Seleccionar CSV de pruebas desde terminal
    print(f"\n Buscando archivos de pruebas para N={n_nodos}...")
    ruta_csv = _seleccionar_csv_pruebas(n_nodos)
    print(f" CSV seleccionado: {ruta_csv.name}")

    subcarpeta = "con_estado_vacio" if permitir_vacio else "sin_estado_vacio"
    results_block_dir = QNODES_ROOT / "results" / "block" / subcarpeta
    results_block_dir.mkdir(parents=True, exist_ok=True)
    fecha_actual = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    nombre_excel = f"results_KQNodes_N{n_nodos}_k{k_nombre}_{fecha_actual}.xlsx"
    ruta_salida = results_block_dir / nombre_excel

    # Leer CSV
    with open(ruta_csv, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        pruebas = list(reader)

    if not pruebas:
        print(" Error: El CSV de pruebas está vacío.")
        sys.exit(1)

    def _get_col(row: dict, *nombres: str) -> str:
        for nombre in nombres:
            for key, val in row.items():
                if key.strip().lower() == nombre.lower():
                    return str(val).strip() if val is not None else ""
        return ""

    # Etiquetar cada prueba con su posición original antes de reordenar
    for orig_idx, row in enumerate(pruebas):
        row["_orig_idx"] = orig_idx

    pruebas.sort(key=lambda row: (
        len(_get_col(row, "alcance o purview (t+1)", "alcance", "purview", "effect")),
        len(_get_col(row, "mecanismo(t)", "mecanismo", "mechanism", "cause")),
    ))
    print(" Pruebas reordenadas de menor a mayor complejidad (por nodos en alcance).")

    sep(f" Iniciando lote: {len(pruebas)} prueba(s) — '{ruta_csv.name}'")
    print(f"   Dataset TPM     : {ruta_tpm.name}")
    print(f"   Sist. Candidato : {bits_candidato}  ({n_candidatos} nodos)")
    print(f"   Estado Inicial  : {estado}")
    print(f"   K global        : {k_nombre}")
    print(f"   Presente vacío  : {'Sí' if permitir_vacio else 'No'}")
    print(f"   Salida          : {ruta_salida}\n")

    # Crear Excel con filas en "pendiente..." y fila de tiempo total al final
    pruebas_para_excel = sorted(
        [
            {
                "_orig_idx":      row["_orig_idx"],
                "#prueba":        _get_col(row, "#prueba", "prueba", "num", "n") or str(row["_orig_idx"] + 1),
                "alcance_orig":   _get_col(row, "alcance o purview (t+1)", "alcance", "purview", "effect"),
                "mecanismo_orig": _get_col(row, "mecanismo(t)", "mecanismo", "mechanism", "cause"),
            }
            for row in pruebas
        ],
        key=lambda r: r["_orig_idx"],
    )
    _crear_excel_bloque_inicial(ruta_salida, pruebas_para_excel, len(pruebas))
    fila_total = len(pruebas) + 2
    fila_warm  = fila_total + 1

    print(f" Excel creado ({len(pruebas_para_excel)} filas). Los resultados se guardarán prueba a prueba.")
    print(f"   {ruta_salida}\n")

    # Deshabilitar profiling para el lote (evita crear cientos de HTMLs)
    gestor_perfilado.enabled = False

    # Crear instancia una sola vez — aplicar_estrategia limpia cachés internos
    analizador = DynamicPartition(tpm)
    resultados = []
    total = len(pruebas)
    tiempo_inicio_lote = time.time()
    # Acumulado del tiempo de "calentar motores" (preparación del subsistema),
    # aparte del tiempo de las pruebas. Tras la primera prueba las cachés de
    # condicionado hacen que las siguientes preparaciones sean casi nulas.
    tiempo_warmup_acum = 0.0

    for idx, row in enumerate(pruebas, start=1):
        num_prueba   = _get_col(row, "#prueba", "prueba", "num", "n") or str(idx)
        alcance_raw  = _get_col(row, "alcance o purview (t+1)", "alcance", "purview", "effect")
        mecanism_raw = _get_col(row, "mecanismo(t)", "mecanismo", "mechanism", "cause")

        orig_idx = row.get("_orig_idx", idx - 1)
        print(f" Prueba #{num_prueba:<4} ({idx}/{total})", end="", flush=True)

        if not alcance_raw or not mecanism_raw:
            msg = "Alcance o mecanismo vacíos en el CSV"
            print(f"  ✗ {msg}")
            _fila = {
                "_orig_idx": orig_idx,
                "#prueba": num_prueba,
                "alcance_orig": alcance_raw,
                "mecanismo_orig": mecanism_raw,
                "error": msg,
            }
            resultados.append(_fila)
            _actualizar_fila_excel(ruta_salida, orig_idx + 2, _fila)
            _actualizar_tiempo_total_excel(ruta_salida, fila_total, time.time() - tiempo_inicio_lote)
            continue

        bits_alcance   = _letras_a_binario(alcance_raw, n_nodos)
        bits_mecanismo = _letras_a_binario(mecanism_raw, n_nodos)

        invalidos_alc = [
            i for i, (a, c) in enumerate(zip(bits_alcance, bits_candidato))
            if a == "1" and c == "0"
        ]
        if invalidos_alc:
            msg = f"Alcance activa nodos {invalidos_alc} fuera del candidato"
            print(f"  ✗ {msg}")
            _fila = {"_orig_idx": orig_idx, "#prueba": num_prueba,
                     "alcance_orig": alcance_raw, "mecanismo_orig": mecanism_raw, "error": msg}
            resultados.append(_fila)
            _actualizar_fila_excel(ruta_salida, orig_idx + 2, _fila)
            _actualizar_tiempo_total_excel(ruta_salida, fila_total, time.time() - tiempo_inicio_lote)
            continue

        invalidos_mec = [
            i for i, (m, c) in enumerate(zip(bits_mecanismo, bits_candidato))
            if m == "1" and c == "0"
        ]
        if invalidos_mec:
            msg = f"Mecanismo activa nodos {invalidos_mec} fuera del candidato"
            print(f"  ✗ {msg}")
            _fila = {"_orig_idx": orig_idx, "#prueba": num_prueba,
                     "alcance_orig": alcance_raw, "mecanismo_orig": mecanism_raw, "error": msg}
            resultados.append(_fila)
            _actualizar_fila_excel(ruta_salida, orig_idx + 2, _fila)
            _actualizar_tiempo_total_excel(ruta_salida, fila_total, time.time() - tiempo_inicio_lote)
            continue

        try:
            solucion = analizador.aplicar_estrategia(
                estado_inicial=estado,
                condicion=bits_candidato,
                alcance=bits_alcance,
                mecanismo=bits_mecanismo,
                k=k_val_global,
                permitir_presente_vacio=permitir_vacio,
            )
            tiempo_seg = float(solucion.tiempo_ejecucion)
            tiempo_fmt = formatear_tiempo(tiempo_seg)
            tiempo_warmup_acum += float(getattr(solucion, "tiempo_preparacion", 0.0))

            print(f"  ✓  {tiempo_fmt}")

            _fila = {
                "_orig_idx": orig_idx,
                "#prueba": num_prueba,
                "alcance_orig": alcance_raw,
                "mecanismo_orig": mecanism_raw,
                "perdida_phi": float(solucion.perdida),
                "particion": str(solucion.particion),
                "tiempo_formateado": tiempo_fmt,
                "error": None,
            }
            resultados.append(_fila)
            _actualizar_fila_excel(ruta_salida, orig_idx + 2, _fila)
            _actualizar_tiempo_total_excel(ruta_salida, fila_total, time.time() - tiempo_inicio_lote)
            _actualizar_warmup_excel(ruta_salida, fila_warm, tiempo_warmup_acum)

        except Exception as e:
            msg = str(e)
            print(f"  ✗  Error: {msg}")
            _fila = {
                "_orig_idx": orig_idx,
                "#prueba": num_prueba,
                "alcance_orig": alcance_raw,
                "mecanismo_orig": mecanism_raw,
                "error": msg,
            }
            resultados.append(_fila)
            _actualizar_fila_excel(ruta_salida, orig_idx + 2, _fila)
            _actualizar_tiempo_total_excel(ruta_salida, fila_total, time.time() - tiempo_inicio_lote)

    tiempo_total_lote = time.time() - tiempo_inicio_lote
    _actualizar_tiempo_total_excel(ruta_salida, fila_total, tiempo_total_lote)
    _actualizar_warmup_excel(ruta_salida, fila_warm, tiempo_warmup_acum)

    exitosas  = sum(1 for r in resultados if r.get("error") is None)
    con_error = len(resultados) - exitosas

    sep(" Lote Completado")
    print(f"   Total pruebas   : {len(pruebas)}")
    print(f"   Exitosas        : {exitosas}")
    print(f"   Con error       : {con_error}")
    print(f"   Tiempo total    : {formatear_tiempo(tiempo_total_lote)}")
    print(f"   Arranque motor  : {formatear_tiempo(tiempo_warmup_acum)}")
    print(f"\n ✓ Resultados guardados en: {ruta_salida}")


# ── Punto de entrada ───────────────────────────────────────────────────────

def main():
    aplicacion.desactivar_profiling()

    sep(" KQNodes — k-Partición de Mínima Información (Óptimo Global)")

    opcion = input(
        "\n ¿Qué modo desea ejecutar?\n"
        "   1. Manual     — ingresar una prueba de forma interactiva\n"
        "   2. Por bloque — cargar lote de pruebas desde un archivo CSV\n"
        " Seleccione (1 o 2): "
    ).strip()

    print()
    ruta_archivo, tpm, n_nodos = _seleccionar_tpm()
    print(f" TPM cargada: {ruta_archivo.name}  ({n_nodos} nodos)\n")

    if opcion == "2":
        modo_bloque(ruta_archivo, tpm, n_nodos)
    else:
        sep(" Ejecución Manual — KQNodes Óptimo Global")
        modo_manual(ruta_archivo, tpm, n_nodos)


if __name__ == '__main__':
    main()
