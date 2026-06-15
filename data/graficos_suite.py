# -*- coding: utf-8 -*-
"""
Generación de gráficas comparativas (QNodes vs Geométrica) para la suite 2026-1.

Se invoca al final de run_suite_2026.py sobre el libro ya diligenciado. Lee TODAS
las hojas N disponibles y construye UNA HOJA POR k ("k = 2", "k = 3", "k = 4",
"k = 5"). Cada hoja es autocontenida: lleva su tabla de datos auxiliar a la
izquierda y, a la derecha, sus tres gráficas nativas de Excel. Así la información
de cada tamaño de partición se entiende por separado.

Tres familias de comparaciones por hoja (sobre resultados de K-particiones, k=2,3,4,5):

  (a) Tiempo de ejecución vs TAMAÑO del subsistema (nº de nodos activos del
      alcance ≈ N), GLOBAL (todas las redes juntas), un gráfico por k, series
      QNodes vs Geométrica como nube de puntos + línea de tendencia (polinómica
      orden 2) por motor para ver cómo crece el tiempo al aumentar N.
  (b) Pérdida (Φ) por prueba: barras agrupadas QNodes (azul) vs Geométrica
      (naranja), eje X = nº de prueba, un gráfico por k. La hoja además trae el
      conteo de "¿se halló la misma k-partición?" (comparando las particiones
      normalizadas de ambos motores).
  (c) Dispersión (estilo k-means) Φ QNodes (eje X) vs Φ Geométrica (eje Y), un
      punto por subsistema y un gráfico por k. Los puntos sobre la diagonal y=x
      son donde ambos motores coinciden; la distancia a la diagonal mide cuánta
      diferencia hay entre QNodes y Geométrica.

Los tiempos en el libro están como texto formateado ("0.0865 s", "1 min 48.70 s",
"2 h 3 min 4 s"); aquí se parsean a segundos para poder graficarlos.
"""

import re
from typing import Optional

from openpyxl.chart import ScatterChart, BarChart, LineChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

KS = (2, 3, 4, 5)

_HDR_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center")


# ── Parseo de tiempo formateado → segundos ──────────────────────────────────

def parse_tiempo(valor) -> Optional[float]:
    """Convierte '0.0865 s' / '1 min 48.70 s' / '2 h 3 min 4 s' a segundos (float)."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    # Normalizamos la coma decimal (locale español) a punto: las filas de k=2
    # se diligenciaron a mano y guardan tiempos como '0,1674' / '226,7341' (s),
    # a veces sin sufijo 's'. Las columnas de tiempo nunca llevan separador de
    # miles, así que reemplazar ',' por '.' es seguro aquí.
    s = str(valor).strip().replace(",", ".")
    if not s:
        return None
    h = re.search(r"([\d.]+)\s*h", s)
    m = re.search(r"([\d.]+)\s*min", s)
    seg = re.search(r"([\d.]+)\s*s(?!\w)", s)
    total = 0.0
    encontrado = False
    if h:
        total += float(h.group(1)) * 3600.0; encontrado = True
    if m:
        total += float(m.group(1)) * 60.0; encontrado = True
    if seg:
        total += float(seg.group(1)); encontrado = True
    if encontrado:
        return total
    # Sin unidades h/min/s: número suelto interpretado como segundos.
    try:
        return float(s)
    except ValueError:
        return None


# ── Normalización de particiones (para comparar entre motores) ───────────────

def _canon_particion(texto) -> Optional[frozenset]:
    """
    Lleva una partición (texto de 2 líneas) a forma canónica comparable entre
    motores: frozenset de (frozenset futuros, frozenset presentes), en MAYÚSCULAS.

    Funciona para ambos formatos: QNodes (⎛ ⎞⎝ ⎠) y Geométrica (| |), porque ambos
    ponen futuros en la línea 1 y presentes en la línea 2.
    """
    if not texto:
        return None
    s = str(texto)
    if s.strip().upper().startswith("ERROR"):
        return None
    lineas = s.split("\n")
    if len(lineas) < 2:
        lineas = lineas + [""]

    def celdas(linea: str):
        t = re.sub(r"[⎛⎝⎞⎠]", "|", linea)
        return [c.strip() for c in t.split("|") if c.strip() != ""]

    fut = celdas(lineas[0])
    pres = celdas(lineas[1])
    if not fut:
        return None

    partes = []
    for i, fcell in enumerate(fut):
        fset = (frozenset() if fcell == "∅"
                else frozenset(tok.strip().upper() for tok in fcell.split(",") if tok.strip()))
        pcell = pres[i] if i < len(pres) else "∅"
        pset = (frozenset() if pcell == "∅"
                else frozenset(tok.strip().upper() for tok in pcell.split(",") if tok.strip()))
        partes.append((fset, pset))
    return frozenset(partes)


def _num(valor) -> Optional[float]:
    """Devuelve float si el valor es numérico (no ERROR/None/texto).

    Acepta también números en texto con coma decimal ('0,47558594'), porque las
    pérdidas de k=2 se diligenciaron a mano en locale español. Cualquier texto no
    numérico (p. ej. 'ERROR: ...') devuelve None.
    """
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        t = valor.strip().replace(",", ".")
        if not t:
            return None
        try:
            return float(t)
        except ValueError:
            return None
    return None


# ── Recolección de registros desde las hojas N ───────────────────────────────

def _recolectar(wb, plan_n, cols) -> "list[dict]":
    """
    Recorre las hojas N disponibles y junta un registro por (hoja, prueba, k) con
    pérdida/tiempo de ambos motores. Solo incluye registros con datos completos.
    """
    registros = []
    for n, (hoja, _tpm) in plan_n.items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        fila = 6
        while True:
            alc = ws.cell(fila, 2).value
            if alc is None or str(alc).strip() == "":
                break  # fin de pruebas (las filas de resumen tienen alcance vacío)
            alc_lbl = str(alc).strip()
            size = len(set(alc_lbl.upper()))
            mec = ws.cell(fila, 3).value
            mec_lbl = str(mec).strip() if mec is not None else ""
            # Las pruebas empiezan en la fila 6 y la col 1 es =ROW(A1)=1, =ROW(A2)=2, …
            # ⇒ el nº de prueba es fila-5 (no depende de leer la fórmula de la celda).
            prueba = fila - 5
            for k in KS:
                c_qn = cols[k]["qnodes"]
                c_ge = cols[k]["geomip"]
                qn_loss = _num(ws.cell(fila, c_qn[1]).value)
                ge_loss = _num(ws.cell(fila, c_ge[1]).value)
                qn_t = parse_tiempo(ws.cell(fila, c_qn[2]).value)
                ge_t = parse_tiempo(ws.cell(fila, c_ge[2]).value)
                if None in (qn_loss, ge_loss, qn_t, ge_t):
                    continue
                qn_part_txt = ws.cell(fila, c_qn[0]).value
                ge_part_txt = ws.cell(fila, c_ge[0]).value
                qn_part = _canon_particion(qn_part_txt)
                ge_part = _canon_particion(ge_part_txt)
                misma = (qn_part is not None and ge_part is not None and qn_part == ge_part)
                registros.append({
                    "n": n, "prueba": prueba, "alcance": alc_lbl, "mecanismo": mec_lbl,
                    "size": size, "k": k,
                    "qn_loss": qn_loss, "ge_loss": ge_loss,
                    "qn_t": qn_t, "ge_t": ge_t,
                    "delta": ge_loss - qn_loss, "misma": misma,
                    "qn_part_txt": str(qn_part_txt) if qn_part_txt is not None else "",
                    "ge_part_txt": str(ge_part_txt) if ge_part_txt is not None else "",
                })
            fila += 1
    return registros


# ── Construcción de hojas + gráficas ─────────────────────────────────────────

def _hoja_k(k) -> str:
    """Nombre de la hoja dedicada a un k (p. ej. 'k = 3')."""
    return f"k = {k}"


def _escribir_encabezado(ws, fila, valores):
    for j, v in enumerate(valores, 1):
        c = ws.cell(fila, j, v)
        c.font = _BOLD
        c.fill = _HDR_FILL
        c.alignment = _CENTER


def generar_graficos(wb, plan_n, cols, log=print) -> bool:
    """
    Construye UNA HOJA POR k ('k = 2' … 'k = 5') en el libro `wb`. Cada hoja
    lleva su tabla de datos a la izquierda y sus tres gráficas a la derecha.

    Devuelve True si generó gráficas; False si no había datos.
    """
    registros = _recolectar(wb, plan_n, cols)
    if not registros:
        log("  [gráficos] no hay datos de K-particiones para graficar — se omite.")
        return False

    # Hojas limpias (idempotente ante re-ejecuciones): borra las hojas por-k y las
    # del layout antiguo ('Datos_Graficos'/'Graficos') si quedaron de corridas previas.
    for nombre in ("Graficos", "Datos_Graficos", *(_hoja_k(k) for k in KS)):
        if nombre in wb.sheetnames:
            del wb[nombre]

    # Columnas de la tabla auxiliar (1-indexado), idénticas en cada hoja k:
    #  1 # (nº de prueba) | 2 N | 3 Alcance | 4 Mecanismo | 5 Tamaño |
    #  6 Φ QNodes | 7 Φ Geom | 8 t QNodes | 9 t Geom | 10 Δ Φ | 11 ¿Misma? |
    # 12 Partición QNodes | 13 Partición Geom
    HEADER = ["#", "N", "Alcance", "Mecanismo", "Tamaño", "Φ QNodes", "Φ Geom",
              "t QNodes (s)", "t Geom (s)", "Δ Φ (Geom−QN)", "¿Misma part?",
              "Partición QNodes", "Partición Geom"]
    NCOL = len(HEADER)
    _PART_WRAP = Alignment(wrap_text=True, vertical="top")
    CHART_COL = "O"  # las gráficas van a la derecha de la tabla (cols A..M)

    hojas = 0
    for k in KS:
        # Orden por nº de prueba ascendente (eje X de la gráfica (b)); a igual
        # prueba, por N. Así el eje X de "Pérdida por prueba" queda ordenado.
        regs_k = sorted([r for r in registros if r["k"] == k],
                        key=lambda r: (r["prueba"], r["n"]))
        if not regs_k:
            continue

        ws = wb.create_sheet(_hoja_k(k))
        ws.sheet_properties.tabColor = "D9E1F2"

        # Título + resumen "¿misma partición?" propio de este k.
        ws.cell(1, 1, f"k = {k} — Comparativa QNodes vs Geométrica").font = Font(bold=True, size=14)
        iguales = sum(1 for r in regs_k if r["misma"])
        distintas = len(regs_k) - iguales
        pct_ig = round(100.0 * iguales / len(regs_k), 1)
        ws.cell(2, 1, f"¿Misma k-partición?  iguales: {iguales} · distintas: {distintas} "
                      f"· {pct_ig}% iguales  (sobre {len(regs_k)} subsistemas)").font = Font(italic=True)

        header_row = 4
        _escribir_encabezado(ws, header_row, HEADER)
        data_start = header_row + 1
        fila = data_start
        for r in regs_k:
            ws.cell(fila, 1, r["prueba"])
            ws.cell(fila, 2, r["n"])
            ws.cell(fila, 3, r["alcance"])
            ws.cell(fila, 4, r["mecanismo"])
            ws.cell(fila, 5, r["size"])
            ws.cell(fila, 6, round(r["qn_loss"], 6))
            ws.cell(fila, 7, round(r["ge_loss"], 6))
            ws.cell(fila, 8, round(r["qn_t"], 4))
            ws.cell(fila, 9, round(r["ge_t"], 4))
            ws.cell(fila, 10, round(r["delta"], 6))
            ws.cell(fila, 11, "Sí" if r["misma"] else "No")
            ws.cell(fila, 12, r["qn_part_txt"]).alignment = _PART_WRAP
            ws.cell(fila, 13, r["ge_part_txt"]).alignment = _PART_WRAP
            fila += 1
        data_end = fila - 1
        ds, de = data_start, data_end

        # Tabla de Excel → botones de filtro por columna (incl. N). Al filtrar N,
        # Excel oculta filas y las gráficas (que grafican solo celdas visibles)
        # se actualizan solas. El displayName debe ser único en TODO el libro.
        ref = f"A{header_row}:{get_column_letter(NCOL)}{data_end}"
        tabla = Table(displayName=f"Datos_k{k}", ref=ref)
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9", showRowStripes=True, showColumnStripes=False)
        ws.add_table(tabla)

        # Anchos cómodos para texto largo (alcance/mecanismo y particiones).
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions[get_column_letter(12)].width = 22
        ws.column_dimensions[get_column_letter(13)].width = 22

        # ── Gráficas de este k (apiladas a la derecha de la tabla) ───────────
        cats = Reference(ws, min_col=1, min_row=ds, max_row=de)

        # (a) Tiempo vs tamaño — scatter, X=tamaño, series QNodes/Geom, con línea
        #     de tendencia por motor para ver cómo crece el tiempo al aumentar N.
        sc = ScatterChart()
        sc.title = f"(a) Tiempo vs tamaño de subsistema — k={k}"
        sc.x_axis.title = "Tamaño del subsistema (nodos del alcance ≈ N)"
        sc.y_axis.title = "Tiempo por prueba: prep+búsqueda (s)"
        sc.height = 9; sc.width = 17
        xref = Reference(ws, min_col=5, min_row=ds, max_row=de)
        for col, nombre, color in ((8, "QNodes", "1F77B4"), (9, "Geométrica", "ED7D31")):
            yref = Reference(ws, min_col=col, min_row=ds, max_row=de)
            serie = Series(yref, xref, title=nombre)
            serie.marker = Marker(symbol="circle", size=5)
            serie.graphicalProperties.line.noFill = True  # solo puntos
            # Tendencia exponencial (el tiempo crece de forma exponencial con N);
            # sin ecuación ni R² en el gráfico, solo la curva.
            serie.trendline = Trendline(trendlineType="exp")
            sc.series.append(serie)
        ws.add_chart(sc, f"{CHART_COL}1")

        # (b) Pérdida Φ por prueba — barras: QNodes (azul) vs Geom (naranja).
        bc = BarChart()
        bc.type = "col"
        bc.grouping = "clustered"
        bc.title = f"(b) Pérdida Φ por prueba — k={k}"
        bc.x_axis.title = "# de prueba"
        bc.y_axis.title = "Pérdida Φ (EMD)"
        bc.height = 9; bc.width = 17
        datos_b = Reference(ws, min_col=6, max_col=7, min_row=header_row, max_row=de)
        bc.add_data(datos_b, titles_from_data=True)
        bc.set_categories(cats)
        # Serie 0 = Φ QNodes (azul), serie 1 = Φ Geom (naranja).
        for serie, color in zip(bc.series, ("1F77B4", "ED7D31")):
            serie.graphicalProperties.solidFill = color
        ws.add_chart(bc, f"{CHART_COL}20")

        # (c) Dispersión Φ QNodes vs Φ Geom — un punto por subsistema. Los puntos
        #     sobre la diagonal y=x son donde ambos motores coinciden; cuanto más
        #     se alejan, mayor la diferencia entre QNodes y Geométrica.
        disp = ScatterChart()
        disp.title = f"(c) Dispersión Φ: QNodes vs Geométrica — k={k}"
        disp.x_axis.title = "Φ QNodes"
        disp.y_axis.title = "Φ Geométrica"
        disp.height = 9; disp.width = 17
        x_qn = Reference(ws, min_col=6, min_row=ds, max_row=de)
        y_ge = Reference(ws, min_col=7, min_row=ds, max_row=de)
        s_disp = Series(y_ge, x_qn, title="Φ Geom vs Φ QNodes")
        s_disp.marker = Marker(symbol="circle", size=6)
        s_disp.graphicalProperties.line.noFill = True  # solo puntos (nube k-means)
        disp.series.append(s_disp)
        ws.add_chart(disp, f"{CHART_COL}39")

        hojas += 1

    log(f"  [gráficos] generadas {hojas} hojas (una por k) con tabla + 3 gráficas c/u "
        f"({len(registros)} registros).")
    return True
