# -*- coding: utf-8 -*-
"""
Suite de pruebas 2026-1 — KGeoMIP vs KQNodes para k = 3, 4 y 5.

NO modifica el archivo original: genera una COPIA fechada del libro
DatosPruebas2026_1.xlsx dentro de results_test/ con el nombre
``DatosPruebas2026_1_<fecha>.xlsx`` y escribe ahí todos los resultados.

Recorre en secuencia las hojas N del libro (N10 → N15 → N20 → N22 → N25). Por cada N:
corre KGeoMIP con k=3,4,5 y luego KQNodes con k=3,4,5, sobre todas las pruebas
(Alcance/Mecanismo) que la hoja ya define. Cada lote (N, motor, k) se ejecuta en
un subproceso aislado (_worker_motor.py) porque KGeoMIP y KQNodes comparten el
nombre de paquete `src` y no pueden importarse en el mismo proceso.

Las TPMs y los datos viven en la carpeta data/ de la RAÍZ del repositorio (la
misma que usan KGeoMIP y KQNodes); sólo los resultados van a results_test/.

Por cada prueba terminada se escribe Partición / Pérdida / Tiempo (preparación de su
subsistema + BÚSQUEDA, el costo real de pared por prueba) en la columna que corresponde,
PERO la copia en disco SÓLO se guarda al terminar cada lote (N, motor, k) — guardar por
prueba ralentizaba mucho el suite, sobre todo en N altos. Si una prueba falla se
marca ERROR en memoria y se persiste igual al cerrar el lote; si el proceso se
cae a mitad de un k se pierde lo no guardado de ese k y basta re-ejecutarlo.
Además, al terminar cada lote (N, motor, k) se escriben, debajo de las pruebas y
en la columna de Tiempo de ese motor/k, dos celdas:
  • Tiempo total de las pruebas (Σ de preparación + búsqueda del lote).
  • Justo debajo, el tiempo de "arranque del motor" (warmup): SOLO el boot único del
    subproceso (spawn + import + carga de la TPM + creación del motor). La preparación
    del subsistema por prueba ya va dentro del tiempo de cada prueba, no aquí.

Las pruebas que fallan se marcan como ERROR y se continúa. Las columnas de
BIPARTICIONES nunca se tocan.

Layout de columnas por hoja (1-indexado):
    3-particiones : KQNodes J,K,L (10,11,12)   Geometric M,N,O (13,14,15)
    4-particiones : KQNodes P,Q,R (16,17,18)   Geometric S,T,U (19,20,21)
    5-particiones : KQNodes V,W,X (22,23,24)   Geometric Y,Z,AA (25,26,27)

Uso:
    .venv/Scripts/python.exe data/run_suite_2026.py [opciones]

Opciones:
    --solo-n 10,15        Limita a estos N (por defecto 10,15,20,22,25).
    --solo-k 3,4          Limita a estos k (por defecto 3,4,5).
    --solo-motor qnodes   Limita a un motor (geomip | qnodes; por defecto ambos).
    --vacio / --no-vacio  Permite (o no) mecanismo ∅ en las partes. Por defecto SÍ
                          (permitir_presente_vacio=True), que da la pérdida mínima
                          real; con --no-vacio las pérdidas de KGeoMIP se inflan.
    --rehacer             Recalcula también celdas que ya tienen resultado.
                          (Por defecto se SALTAN las celdas ya rellenadas.)
    --no-interactivo      Usa los CSV predeterminados de PLAN_N sin preguntar.
"""

import os
import sys
import json
import time
import shutil
import argparse
import tempfile
import subprocess
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ── Rutas base ─────────────────────────────────────────────────────────────
REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_SOURCE = REPO_ROOT / "DatosPruebas2026_1.xlsx"
RESULTS_TEST_DIR = REPO_ROOT / "results_test"
WORKER = Path(__file__).resolve().parent / "_worker_motor.py"
PYTHON = sys.executable

# Datos compartidos en la carpeta data/ de la raíz (no por motor).
DATA_ROOT = REPO_ROOT / "data"
SAMPLES_DIR = DATA_ROOT / "samples_binary"

GEOMIP_ROOT = REPO_ROOT / "KGeoMIP"
QNODES_ROOT = REPO_ROOT / "KQNodes"

SAMPLES_DIRS = [SAMPLES_DIR, DATA_ROOT / "samples_no_binary"]

RESULT_SENTINEL = "@@RESULT@@"

# Mapa N → (hoja, archivo TPM dentro de data/samples_binary).
# El sheet de N15 es "15B-Elementos" pero el único TPM binario disponible es N15A.csv.
PLAN_N = {
    10: ("10A-Elementos", "N10A.csv"),
    15: ("15B-Elementos", "N15A.csv"),
    20: ("20A-Elementos", "N20A.csv"),
    22: ("22A-Elementos", "N22A.csv"),
    25: ("25A-Elementos ", "N25A.csv"),  # OJO: el nombre de la hoja lleva un espacio final.
}

# k → motor → (col_particion, col_perdida, col_tiempo)
COLS = {
    3: {"qnodes": (10, 11, 12), "geomip": (13, 14, 15)},
    4: {"qnodes": (16, 17, 18), "geomip": (19, 20, 21)},
    5: {"qnodes": (22, 23, 24), "geomip": (25, 26, 27)},
}

ENGINE_ROOTS = {"geomip": GEOMIP_ROOT, "qnodes": QNODES_ROOT}
ENGINE_NOMBRE = {"geomip": "KGeoMIP", "qnodes": "KQNodes"}
# Orden de motores por N: primero KGeoMIP, luego KQNodes (según lo pedido).
ENGINE_ORDER = ["geomip", "qnodes"]

_PART_ALIGN = Alignment(wrap_text=True, vertical="top", horizontal="left")
_CENTER = Alignment(horizontal="center", vertical="center")
_ERR_FONT = Font(color="CC0000")
_TOT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_WARM_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
_BOLD = Font(bold=True)


# ── Utilidades ─────────────────────────────────────────────────────────────

def formatear_tiempo(segundos: float) -> str:
    if segundos < 60:
        return f"{segundos:.4f} s"
    if segundos < 3600:
        return f"{int(segundos // 60)} min {segundos % 60:.2f} s"
    horas = int(segundos // 3600)
    resto = segundos % 3600
    return f"{horas} h {int(resto // 60)} min {resto % 60:.2f} s"


def etiquetas_a_bits(etiquetas: str, n: int) -> str:
    """Convierte 'ABDEG' a máscara binaria de longitud n. Válido para N<=26 (1 letra/nodo)."""
    activos = set(str(etiquetas).strip().upper())
    return "".join("1" if chr(65 + i) in activos else "0" for i in range(n))


def celda_vacia(ws, fila: int, col: int) -> bool:
    """True si la celda de Partición está vacía / en pendiente (no es un resultado real)."""
    v = ws.cell(row=fila, column=col).value
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s == "pendiente..." or s.startswith("ERROR")


def leer_pruebas(ws, n: int) -> list:
    """Extrae [(fila, alcance_bits, mecanismo_bits, alcance_lbl, mecanismo_lbl)] desde la fila 6."""
    pruebas = []
    fila = 6
    while True:
        alc = ws.cell(row=fila, column=2).value
        mec = ws.cell(row=fila, column=3).value
        if (alc is None or str(alc).strip() == "") and (mec is None or str(mec).strip() == ""):
            break
        alc_lbl = str(alc).strip() if alc else ""
        mec_lbl = str(mec).strip() if mec else ""
        if alc_lbl and mec_lbl:
            pruebas.append((
                fila,
                etiquetas_a_bits(alc_lbl, n),
                etiquetas_a_bits(mec_lbl, n),
                alc_lbl, mec_lbl,
            ))
        fila += 1
    return pruebas


def escribir_resultado(ws, fila: int, cols, res: dict) -> None:
    """Escribe Partición/Pérdida/Tiempo (sólo búsqueda) o ERROR en la fila/columnas dadas."""
    c_part, c_perd, c_tiem = cols
    if res.get("ok"):
        cp = ws.cell(row=fila, column=c_part, value=res["particion"])
        cp.alignment = _PART_ALIGN
        cp.font = Font()
        ws.cell(row=fila, column=c_perd, value=round(float(res["perdida"]), 6)).alignment = _CENTER
        # Tiempo de la prueba = preparación de su subsistema + búsqueda (costo real de
        # pared por prueba; la preparación es trabajo específico de cada prueba).
        t_prueba = float(res["tiempo"]) + float(res.get("prep", 0.0))
        ws.cell(row=fila, column=c_tiem, value=formatear_tiempo(t_prueba)).alignment = _CENTER
    else:
        cp = ws.cell(row=fila, column=c_part, value=f"ERROR: {res.get('error', 'desconocido')}")
        cp.alignment = _PART_ALIGN
        cp.font = _ERR_FONT
        ws.cell(row=fila, column=c_perd, value=None).alignment = _CENTER
        ws.cell(row=fila, column=c_tiem, value=None).alignment = _CENTER


def escribir_resumen_tiempos(ws, col_tiempo, fila_tests, fila_warm, t_tests, t_warm):
    """Escribe el Σ de tiempos de pruebas y, debajo, el tiempo de arranque (warmup)."""
    ct = ws.cell(row=fila_tests, column=col_tiempo, value=formatear_tiempo(t_tests))
    ct.font = _BOLD
    ct.fill = _TOT_FILL
    ct.alignment = _CENTER
    cw = ws.cell(row=fila_warm, column=col_tiempo, value=formatear_tiempo(t_warm))
    cw.font = _BOLD
    cw.fill = _WARM_FILL
    cw.alignment = _CENTER


def escribir_etiquetas_resumen(ws, fila_tests, fila_warm):
    """Etiqueta (col 1) las dos filas de resumen de tiempos de una hoja."""
    et = ws.cell(row=fila_tests, column=1, value="Tiempo total pruebas (Σ prep+búsqueda)")
    et.font = _BOLD
    et.fill = _TOT_FILL
    et.alignment = _CENTER
    ew = ws.cell(row=fila_warm, column=1, value="Arranque motor (warmup)")
    ew.font = _BOLD
    ew.fill = _WARM_FILL
    ew.alignment = _CENTER


# ── Selección interactiva de CSV por N ─────────────────────────────────────

def buscar_csvs_disponibles(n: int) -> list:
    """Devuelve los Path de los CSV para N{n} encontrados en las carpetas de samples."""
    resultado = []
    for d in SAMPLES_DIRS:
        if d.exists():
            resultado.extend(sorted(d.glob(f"N{n}*.csv")))
    return resultado


def seleccionar_csvs_interactivo(ns: list) -> dict:
    """
    Para cada N pedido pregunta qué CSV usar o si se omite.
    Retorna {n: (hoja, tpm_path: Path)} sólo para los N que el usuario no omite.
    """
    print("\n" + "=" * 64)
    print("  Selección de TPM (CSV) para cada N")
    print("  Enter sin valor = usar el predeterminado  |  0 = omitir N")
    print("=" * 64)

    plan = {}
    for n in ns:
        if n not in PLAN_N:
            print(f"\nN{n}: sin hoja definida en PLAN_N — se omitirá.")
            continue
        hoja, default_name = PLAN_N[n]
        disponibles = buscar_csvs_disponibles(n)

        print(f"\nN{n}  (hoja: '{hoja}')")
        if not disponibles:
            print("  No se encontraron CSV para este N en las carpetas de samples.")
            resp = input("  Ruta manual al CSV (o Enter para omitir): ").strip()
            if resp:
                p = Path(resp)
                if p.exists():
                    plan[n] = (hoja, p)
                    print(f"  → Usando: {p.name}")
                else:
                    print(f"  Archivo no encontrado — se omite N{n}.")
            else:
                print(f"  Se omite N{n}.")
            continue

        for i, f in enumerate(disponibles, 1):
            marca = "  ◄ predeterminado" if f.name == default_name else ""
            print(f"    [{i}] {f.name}  ({f.parent.name}){marca}")
        print(f"    [0] Omitir N{n}")

        while True:
            resp = input(f"  Selección [1-{len(disponibles)} / 0 / Enter=predeterminado]: ").strip()
            if resp == "0":
                print(f"  N{n} omitido.")
                break
            if resp == "":
                default_path = SAMPLES_DIR / default_name
                if default_path.exists():
                    plan[n] = (hoja, default_path)
                    print(f"  → Usando predeterminado: {default_name}")
                    break
                # predeterminado no existe → buscar en otras carpetas
                alt = next((f for f in disponibles if f.name == default_name), None)
                if alt:
                    plan[n] = (hoja, alt)
                    print(f"  → Usando predeterminado: {alt}")
                    break
                print(f"  El predeterminado '{default_name}' no existe. Elige una opción o 0 para omitir.")
                continue
            try:
                idx = int(resp)
                if 1 <= idx <= len(disponibles):
                    elegido = disponibles[idx - 1]
                    plan[n] = (hoja, elegido)
                    print(f"  → Usando: {elegido.name}")
                    break
                print(f"  Ingresa un número entre 0 y {len(disponibles)}.")
            except ValueError:
                print(f"  Ingresa un número entre 0 y {len(disponibles)}.")

    return plan


# ── Ejecución de un lote (N, motor, k) ─────────────────────────────────────

def correr_lote(wb, ws, dest_path, n, tpm_path, engine, k, estado, candidato_bits,
                pruebas, rehacer, permitir_vacio, fila_tests, fila_warm) -> None:
    cols = COLS[k][engine]
    col_tiempo = cols[2]
    root = ENGINE_ROOTS[engine]

    # Filtrar pruebas pendientes (salvo --rehacer).
    pendientes = [p for p in pruebas if rehacer or celda_vacia(ws, p[0], cols[0])]
    etiqueta = f"N{n} · {ENGINE_NOMBRE[engine]} · k={k}"
    if not pendientes:
        print(f"  [{etiqueta}] todas las celdas ya están listas — se omite.")
        return
    if not tpm_path.exists():
        print(f"  [{etiqueta}] TPM no encontrada: {tpm_path} — se omite.")
        return

    cfg = {
        "engine": engine,
        "root": str(root),
        "tpm": str(tpm_path),
        "estado": estado,
        "candidato": candidato_bits,
        "permitir_presente_vacio": permitir_vacio,
        "k": k,
        "tests": [{"row": fila, "alcance": ab, "mecanismo": mb}
                  for (fila, ab, mb, _, _) in pendientes],
    }

    with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as tf:
        json.dump(cfg, tf, ensure_ascii=False)
        lote_path = tf.name

    total = len(pendientes)
    print(f"  [{etiqueta}] {total} prueba(s) pendientes — lanzando worker...")
    t0 = time.time()
    boot_time = 0.0            # tiempo de spawn + import + creación del motor
    t_tests = 0.0             # Σ de tiempos de búsqueda de las pruebas
    t_prep = 0.0              # Σ de tiempos de preparación de subsistema
    hechas = 0
    try:
        proc = subprocess.Popen(
            [PYTHON, str(WORKER), lote_path],
            stdout=subprocess.PIPE, stderr=subprocess.DEVNULL,
            text=True, encoding="utf-8", bufsize=1,
        )
        for linea in proc.stdout:
            if not linea.startswith(RESULT_SENTINEL):
                continue
            payload = json.loads(linea[len(RESULT_SENTINEL):])
            if "ready" in payload or "fatal" in payload:
                if "ready" in payload:
                    boot_time = time.time() - t0   # boot del motor hasta listo
                if "fatal" in payload:
                    print(f"  [{etiqueta}] ERROR FATAL worker: {payload['fatal']}")
                continue
            fila = payload["row"]
            escribir_resultado(ws, fila, cols, payload)
            if payload.get("ok"):
                # El tiempo de la prueba incluye su preparación de subsistema.
                t_tests += float(payload.get("tiempo", 0.0)) + float(payload.get("prep", 0.0))
                t_prep += float(payload.get("prep", 0.0))   # solo informativo
            # NO se guarda por prueba (lento): el lote completo se persiste al
            # terminar el k (más abajo). Si una prueba falla se marca ERROR en
            # memoria y se persiste igual al cerrar el lote.
            hechas += 1
            estado_txt = "OK " if payload.get("ok") else "ERR"
            print(f"    fila {fila:>4}  {estado_txt}  ({hechas}/{total})  "
                  f"[{formatear_tiempo(time.time() - t0)}]")
        proc.wait()
    finally:
        try:
            os.unlink(lote_path)
        except OSError:
            pass

    # Arranque (warmup) = SOLO el boot único del subproceso: spawn + import + carga de
    # la TPM + creación del motor (hasta "ready"). La preparación de subsistema por
    # prueba ya va incluida en el tiempo de cada prueba (y en t_tests), no aquí.
    t_warmup = boot_time
    escribir_resumen_tiempos(ws, col_tiempo, fila_tests, fila_warm, t_tests, t_warmup)
    wb.save(dest_path)

    print(f"  [{etiqueta}] completado: {hechas}/{total} | pruebas {formatear_tiempo(t_tests)} | "
          f"arranque {formatear_tiempo(t_warmup)} | total {formatear_tiempo(time.time() - t0)}")


# ── Punto de entrada ───────────────────────────────────────────────────────

def main() -> None:
    for _stream in (sys.stdout, sys.stderr):
        try:
            _stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:  # noqa: BLE001
            pass

    ap = argparse.ArgumentParser(description="Suite KGeoMIP/KQNodes k=3,4,5 → results_test/DatosPruebas2026_1_<fecha>.xlsx")
    ap.add_argument("--solo-n", default="10,15,20,22,25")
    ap.add_argument("--solo-k", default="3,4,5")
    ap.add_argument("--solo-motor", default="geomip,qnodes")
    ap.add_argument("--rehacer", action="store_true")
    ap.add_argument("--no-interactivo", dest="interactivo", action="store_false", default=True,
                    help="Usa los CSV predeterminados de PLAN_N sin preguntar.")
    ap.add_argument("--vacio", dest="vacio", action="store_true", default=True,
                    help="Permite mecanismo ∅ (por defecto). Da la pérdida mínima real.")
    ap.add_argument("--no-vacio", dest="vacio", action="store_false",
                    help="Prohíbe mecanismo ∅ (las pérdidas de KGeoMIP se inflan).")
    args = ap.parse_args()

    ns = [int(x) for x in args.solo_n.split(",") if x.strip()]
    ks = [int(x) for x in args.solo_k.split(",") if x.strip()]
    motores = [m.strip().lower() for m in args.solo_motor.split(",") if m.strip()]
    motores = [m for m in ENGINE_ORDER if m in motores]   # respeta orden KGeoMIP→KQNodes

    if not XLSX_SOURCE.exists():
        sys.exit(f"No se encontró {XLSX_SOURCE}")

    # ── Selección interactiva de CSV (o predeterminados) ──────────────────────
    if args.interactivo:
        plan_elegido = seleccionar_csvs_interactivo(ns)
    else:
        plan_elegido = {
            n: (PLAN_N[n][0], SAMPLES_DIR / PLAN_N[n][1])
            for n in ns if n in PLAN_N
        }

    if not plan_elegido:
        sys.exit("No se seleccionó ningún N — nada que ejecutar.")

    # ── Copia fechada en results_test/ — el original NO se toca ──────────────
    RESULTS_TEST_DIR.mkdir(parents=True, exist_ok=True)
    fecha = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    dest_path = RESULTS_TEST_DIR / f"DatosPruebas2026_1_{fecha}.xlsx"
    shutil.copy2(XLSX_SOURCE, dest_path)

    wb = openpyxl.load_workbook(dest_path)
    t_inicio = time.time()

    ns_elegidos = list(plan_elegido.keys())
    print("\n" + "=" * 64)
    print(f"  Suite 2026-1   N={ns_elegidos}   k={ks}   motores={[ENGINE_NOMBRE[m] for m in motores]}")
    print(f"  permitir_presente_vacio = {args.vacio}")
    print(f"  Fuente : {XLSX_SOURCE.name}")
    print(f"  Salida : {dest_path}")
    print("=" * 64)

    for n, (hoja, tpm_path) in plan_elegido.items():
        if hoja not in wb.sheetnames:
            print(f"\nN{n}: hoja '{hoja}' no existe en el libro — se omite.")
            continue
        ws = wb[hoja]

        estado = str(ws.cell(1, 2).value).strip()
        candidato_lbl = str(ws.cell(3, 2).value).strip()
        candidato_bits = etiquetas_a_bits(candidato_lbl, n)
        pruebas = leer_pruebas(ws, n)

        # Filas de resumen de tiempos: dos filas debajo de la última prueba.
        ultima_fila = 5 + len(pruebas)            # las pruebas empiezan en la fila 6
        fila_tests = ultima_fila + 2
        fila_warm = ultima_fila + 3
        escribir_etiquetas_resumen(ws, fila_tests, fila_warm)

        print(f"\n{'─' * 64}\nN{n}  (hoja '{hoja}', TPM {tpm_path.name})")
        print(f"  Estado={estado}  Candidato={candidato_bits}  Pruebas={len(pruebas)}")

        for engine in motores:           # KGeoMIP primero, luego KQNodes
            for k in ks:                  # k = 3, 4, 5
                correr_lote(wb, ws, dest_path, n, tpm_path, engine, k,
                            estado, candidato_bits, pruebas, args.rehacer,
                            args.vacio, fila_tests, fila_warm)

    # ── Gráficas comparativas (QNodes vs Geométrica) ──────────────────────────
    try:
        from graficos_suite import generar_graficos
        print("\nGenerando gráficas comparativas...")
        generar_graficos(wb, PLAN_N, COLS)
    except Exception as exc:  # noqa: BLE001 — las gráficas no deben tumbar el suite
        print(f"  [gráficos] no se pudieron generar: {type(exc).__name__}: {exc}")

    wb.save(dest_path)
    print(f"\n{'=' * 64}")
    print(f"  Suite finalizada en {formatear_tiempo(time.time() - t_inicio)}")
    print(f"  Resultados en: {dest_path}")
    print("=" * 64)


if __name__ == "__main__":
    main()
