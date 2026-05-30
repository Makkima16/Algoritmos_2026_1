# -*- coding: utf-8 -*-
"""
Script de ejecución interactivo para KGeoMIP — k-Partición de Mínima Información (ÓPTIMO GLOBAL)

Modos de ejecución:
  1. Manual   — ingreso interactivo de una sola prueba.
  2. Por bloque — carga un CSV con múltiples pruebas y guarda los resultados
                  en el destino que el usuario indique.

Formato del CSV para modo bloque:
  #Prueba,Alcance o Purview (t+1),Mecanismo(t)
  1,ABCDE,ABCDE
  2,ABCDE,ABCD
  3,ACE,BDF

  - Alcance / Mecanismo: etiquetas de nodos (A=nodo 0, B=nodo 1, ...).
  - k se ingresa globalmente al inicio del lote (o vacío = todas las k).
  - El candidato y el estado inicial se ingresan una sola vez para todo el lote.
  - El resultado se guarda automáticamente en GeoMIP/results/block/ como .xlsx.
"""

import sys
import csv
import logging
import random
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
import numpy as np
import json
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.models.base.application import aplicacion
from src.controllers.manager import Manager
from src.controllers.strategies.kgeomip import KGeoMIP
from src.controllers.strategies.kgeomip import _serializar_particion
from src.funcs.base import ABECEDARY
from src.lazy_tpm import cargar_tpm

GEOMIP_ROOT = Path(__file__).resolve().parents[2]
SAMPLES_DIR = GEOMIP_ROOT / "data" / "samples"


# ── Utilidades de presentación ─────────────────────────────────────────────

def sep(titulo: str):
    print(f"\n{'='*60}")
    print(f"  {titulo}")
    print('='*60)


def formatear_tiempo(segundos: float) -> str:
    """Convierte segundos en cadena legible: s / min s / h min s."""
    if segundos < 60:
        return f"{segundos:.4f} s"
    elif segundos < 3600:
        mins = int(segundos // 60)
        segs = segundos % 60
        return f"{mins} min {segs:.2f} s"
    else:
        horas = int(segundos // 3600)
        resto = segundos % 3600
        mins = int(resto // 60)
        segs = resto % 60
        return f"{horas} h {mins} min {segs:.2f} s"


# ── Helpers de entrada ─────────────────────────────────────────────────────

def _seleccionar_tpm() -> tuple:
    """Abre diálogo para seleccionar la TPM y la carga. Retorna (ruta, tpm, n_nodos)."""
    print(" Abriendo explorador de archivos en la carpeta de datasets...")
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        ruta_str = filedialog.askopenfilename(
            initialdir=str(SAMPLES_DIR),
            title="Seleccionar matriz de probabilidad (TPM) CSV",
            filetypes=(("CSV files", "*.csv"), ("All files", "*.*"))
        )
        root.destroy()
    except Exception as e:
        print(f" No se pudo cargar el explorador: {e}")
        archivo = input(" Ingrese manualmente el nombre del archivo (ej. N3A.csv): ").strip()
        ruta_str = str(SAMPLES_DIR / archivo)

    if not ruta_str:
        print(" Ejecución cancelada: No se seleccionó ningún archivo.")
        sys.exit(1)

    ruta = Path(ruta_str)
    if not ruta.exists():
        print(f" Error: El archivo {ruta} no fue encontrado.")
        sys.exit(1)

    try:
        tpm = cargar_tpm(ruta)
        n_nodos = tpm.shape[1] if len(tpm.shape) > 1 else int(np.log2(tpm.shape[0]))
    except Exception as e:
        print(f" Error al leer el fichero TPM: {e}")
        sys.exit(1)

    return ruta, tpm, n_nodos


def _pedir_candidato(n_nodos: int) -> tuple:
    """Solicita el sistema candidato/condición. Retorna (bits_candidato, n_candidatos)."""
    print(f"\n La red consta de {n_nodos} variables lógicas asociadas.")
    candidato_input = input(
        f" Ingrese el SISTEMA CANDIDATO en binario (longitud: {n_nodos} bits)\n"
        f"   O presione [ENTER] para tomar TODO el sistema: "
    ).strip()

    if not candidato_input:
        bits_candidato = "1" * n_nodos
        print(f" Seleccionado el sistema completo: {bits_candidato}")
    else:
        if set(candidato_input).issubset({"0", "1"}) and len(candidato_input) == n_nodos:
            bits_candidato = candidato_input
        else:
            print(f" Error: La máscara debe ser cadena de unos y ceros de longitud {n_nodos}.")
            sys.exit(1)

    n_candidatos = bits_candidato.count("1")
    if n_candidatos < 2:
        print(" Error: El sistema candidato debe tener al menos 2 variables (1s).")
        sys.exit(1)

    return bits_candidato, n_candidatos


def _reconstruir_estado(estado_candidato: str, bits_candidato: str, n_nodos: int) -> str:
    """Expande el estado del candidato al vector de longitud n_nodos."""
    estado_completo = ["0"] * n_nodos
    idx = 0
    for i, bit in enumerate(bits_candidato):
        if bit == "1":
            estado_completo[i] = estado_candidato[idx]
            idx += 1
    return "".join(estado_completo)


def _pedir_estado(n_candidatos: int, bits_candidato: str, n_nodos: int) -> tuple:
    """Solicita el estado inicial. Retorna (estado_candidato, estado_completo)."""
    estado_input = input(
        f"\n Ingrese el ESTADO INICIAL en binario (longitud: {n_candidatos} bits)\n"
        f"   O presione [ENTER] para uno RANDOM: "
    ).strip()

    if not estado_input:
        estado_candidato = "".join(str(random.randint(0, 1)) for _ in range(n_candidatos))
        print(f" Estado aleatorio auto-generado: {estado_candidato}")
    else:
        if set(estado_input).issubset({"0", "1"}) and len(estado_input) == n_candidatos:
            estado_candidato = estado_input
        else:
            print(" Error: El estado inicial debe ser una cadena de unos y ceros.")
            sys.exit(1)

    estado = _reconstruir_estado(estado_candidato, bits_candidato, n_nodos)
    return estado_candidato, estado


def _guardar_resultado_single(
    ruta_archivo: Path,
    estado: str,
    bits_candidato: str,
    bits_alcance: str,
    bits_mecanismo: str,
    solucion_geomip,
    kgeomip_inst,
):
    """Guarda el resultado de una ejecución manual en la carpeta results/."""
    results_dir = GEOMIP_ROOT / "results"
    results_dir.mkdir(parents=True, exist_ok=True)

    estrateg_str = getattr(solucion_geomip, "estrategia", "kgeomip")
    match_k = re.search(r'K=(\d+)', estrateg_str)
    k_val_str = match_k.group(1) if match_k else "optimal"

    fecha_actual = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    nombre_archivo = f"{fecha_actual}_resultado_{ruta_archivo.stem}_k{k_val_str}_global.json"
    ruta_salida = results_dir / nombre_archivo

    tiempo_seg = float(
        getattr(solucion_geomip, "tiempo_ejecucion",
                getattr(solucion_geomip, "tiempo_total", 0.0))
    )

    res_data = {
        "dataset": ruta_archivo.name,
        "estado_inicial": estado,
        "sistema_candidato": bits_candidato,
        "alcance": bits_alcance,
        "mecanismo": bits_mecanismo,
        "estrategia": estrateg_str,
        "perdida_phi": float(solucion_geomip.perdida),
        "fundamento_eleccion": (
            "Se seleccionó esta k-partición porque presenta la pérdida de información "
            "(EMD) mínima global de todo el abanico evaluado."
        ),
        "distribucion_subsistema": (
            solucion_geomip.distribucion_subsistema.tolist()
            if hasattr(solucion_geomip.distribucion_subsistema, 'tolist')
            else solucion_geomip.distribucion_subsistema
        ),
        "distribucion_particion": (
            solucion_geomip.distribucion_particion.tolist()
            if hasattr(solucion_geomip.distribucion_particion, 'tolist')
            else solucion_geomip.distribucion_particion
        ),
        "particion": str(solucion_geomip.particion),
        "historico_comparaciones": getattr(kgeomip_inst, "historico_particiones", []),
        "tiempo_segundos": tiempo_seg,
        "tiempo_formateado": formatear_tiempo(tiempo_seg),
    }

    with open(ruta_salida, 'w', encoding='utf-8') as f:
        json.dump(res_data, f, indent=4, ensure_ascii=False)

    print(f"\n ✓ Resultados guardados en: {ruta_salida.relative_to(GEOMIP_ROOT)}")


# ── Helpers de modo bloque ─────────────────────────────────────────────────

def _letras_a_binario(etiquetas: str, n_nodos: int) -> str:
    """
    Convierte etiquetas de nodos (ej: 'ABCEFG') a máscara binaria de longitud n_nodos.
    Usa la misma notación de ABECEDARY: A, B, ..., Z, AA, AB, ...
    Parsea con matching greedy para soportar etiquetas multi-caracter (N>26).
    """
    etiquetas = etiquetas.strip().upper()
    activos: set[str] = set()
    labels_validas = ABECEDARY[:n_nodos]
    idx = 0
    while idx < len(etiquetas):
        matched = False
        for length in range(min(3, len(etiquetas) - idx), 0, -1):
            lbl = etiquetas[idx:idx + length]
            if lbl in labels_validas:
                activos.add(lbl)
                idx += length
                matched = True
                break
        if not matched:
            idx += 1
    return "".join("1" if ABECEDARY[i] in activos else "0" for i in range(n_nodos))


def _guardar_excel_bloque(
    ruta_salida: Path,
    resultados: list,
) -> None:
    """Crea el archivo Excel con los resultados del lote."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    COLUMNAS = [
        ("#Prueba",                  10),
        ("Alcance o Purview (t+1)",  28),
        ("Mecanismo(t)",             22),
        ("Particion",                48),
        ("Perdida",                  16),
        ("Tiempo",                   20),
    ]

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for col_idx, (titulo, ancho) in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    center_align = Alignment(horizontal="center", vertical="center")
    part_align   = Alignment(wrap_text=True, vertical="top", horizontal="left")
    error_font   = Font(color="CC0000")
    alt_fill     = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for fila_idx, r in enumerate(resultados, start=2):
        relleno = alt_fill if (fila_idx % 2 == 0) else None

        def _c(col: int, valor, alinear=center_align, fuente=None):
            c = ws.cell(row=fila_idx, column=col, value=valor)
            c.alignment = alinear
            if fuente:
                c.font = fuente
            if relleno:
                c.fill = relleno
            return c

        _c(1, r.get("#prueba"))
        _c(2, r.get("alcance_orig", ""))
        _c(3, r.get("mecanismo_orig", ""))

        if r.get("error"):
            _c(4, f"ERROR: {r['error']}", alinear=part_align, fuente=error_font)
            _c(5, None)
            _c(6, None)
        else:
            _c(4, r.get("particion", ""), alinear=part_align)
            perdida = r.get("perdida_phi")
            _c(5, round(perdida, 6) if perdida is not None else None)
            _c(6, r.get("tiempo_formateado", ""))

        ws.row_dimensions[fila_idx].height = 45

    wb.save(ruta_salida)


# ── Modo 1: Ejecución manual ───────────────────────────────────────────────

def modo_manual(ruta_archivo: Path, tpm: np.ndarray, n_nodos: int):
    """Ejecuta una sola prueba de forma interactiva."""
    bits_candidato, n_candidatos = _pedir_candidato(n_nodos)
    _, estado = _pedir_estado(n_candidatos, bits_candidato, n_nodos)

    # ── Alcance ────────────────────────────────────────────────────────────
    print(f"\n Ingrese el ALCANCE (Purview t+1) en binario (longitud: {n_nodos} bits)")
    print(f"   Referencia — Candidato activo: {bits_candidato}")
    print(f"   1 = observar ese nodo en t+1  |  0 = ignorarlo (marginalizarlo)")
    alcance_input = input(
        f"   O presione [ENTER] para usar TODO el candidato como alcance: "
    ).strip()

    if not alcance_input:
        bits_alcance = bits_candidato
        print(f" Alcance: {bits_alcance} (igual al candidato)")
    else:
        if set(alcance_input).issubset({"0", "1"}) and len(alcance_input) == n_nodos:
            invalidos = [
                i for i, (a, c) in enumerate(zip(alcance_input, bits_candidato))
                if a == "1" and c == "0"
            ]
            if invalidos:
                print(f" Error: El alcance activa nodos {invalidos} que NO están en el candidato.")
                sys.exit(1)
            bits_alcance = alcance_input
        else:
            print(f" Error: Debe ser una cadena de 0s y 1s de longitud {n_nodos}.")
            sys.exit(1)

    # ── Mecanismo ──────────────────────────────────────────────────────────
    print(f"\n Ingrese el MECANISMO (t) en binario (longitud: {n_nodos} bits)")
    print(f"   Referencia — Candidato activo: {bits_candidato}")
    print(f"   1 = fijar ese nodo en su estado inicial  |  0 = marginalizarlo")
    mecanismo_input = input(
        f"   O presione [ENTER] para usar TODO el candidato como mecanismo: "
    ).strip()

    if not mecanismo_input:
        bits_mecanismo = bits_candidato
        print(f" Mecanismo: {bits_mecanismo} (igual al candidato)")
    else:
        if set(mecanismo_input).issubset({"0", "1"}) and len(mecanismo_input) == n_nodos:
            invalidos = [
                i for i, (m, c) in enumerate(zip(mecanismo_input, bits_candidato))
                if m == "1" and c == "0"
            ]
            if invalidos:
                print(f" Error: El mecanismo activa nodos {invalidos} que NO están en el candidato.")
                sys.exit(1)
            bits_mecanismo = mecanismo_input
        else:
            print(f" Error: Debe ser una cadena de 0s y 1s de longitud {n_nodos}.")
            sys.exit(1)

    # ── K ─────────────────────────────────────────────────────────────────
    print(f"\n Número máximo de particiones posibles: {n_candidatos}")
    k_input = input(
        f" Ingrese el valor de K (particiones a evaluar) [2-{n_candidatos}]\n"
        f"   O presione [ENTER] para evaluar TODAS las k posibles: "
    ).strip()

    if not k_input:
        k_a_evaluar = None
        print(f" Se evaluarán TODAS las k posibles (2 a {min(5, n_candidatos)})")
    else:
        try:
            k_val = int(k_input)
            if k_val < 2 or k_val > n_candidatos:
                print(f" Error: K debe estar entre 2 y {n_candidatos}.")
                sys.exit(1)
            k_a_evaluar = k_val
            print(f" Se evaluará K = {k_a_evaluar}")
        except ValueError:
            print(" Error: K debe ser un número entero.")
            sys.exit(1)

    # ── Resumen y ejecución ────────────────────────────────────────────────
    sep(" Iniciando Algoritmo — KGeoMIP Partición Óptima Global")
    print(f"   Dataset Cargado : {ruta_archivo.name}")
    print(f"   Sist. Candidato : {bits_candidato} ({n_candidatos} nodos)")
    print(f"   Estado Central  : {estado}")
    print(f"   Alcance (t+1)   : {bits_alcance}")
    print(f"   Mecanismo (t)   : {bits_mecanismo}")
    if k_a_evaluar:
        print(f"   K a evaluar     : {k_a_evaluar}\n")
    else:
        print(f"   K a evaluar     : Todas (2..{min(5, n_candidatos)})\n")

    manager_inst = Manager(estado)
    kgeomip_inst = KGeoMIP(manager_inst)
    solucion_geomip = kgeomip_inst.aplicar_estrategia(
        condicion=bits_candidato,
        alcance=bits_alcance,
        mecanismo=bits_mecanismo,
        tpm=tpm,
        k=k_a_evaluar
    )

    print(solucion_geomip)
    _guardar_resultado_single(
        ruta_archivo, estado, bits_candidato, bits_alcance, bits_mecanismo,
        solucion_geomip, kgeomip_inst,
    )


# ── Modo 2: Ejecución por bloque ───────────────────────────────────────────

def modo_bloque(ruta_tpm: Path, tpm: np.ndarray, n_nodos: int):
    """Ejecuta un lote de pruebas desde un CSV y guarda resultados en Excel."""
    sep(" Modo Bloque — Ejecución por Lotes desde CSV")

    print(
        "\n Formato esperado del CSV de pruebas:\n"
        "   #Prueba,Alcance o Purview (t+1),Mecanismo(t)\n"
        "   1,ABCDE,ABCDE\n"
        "   2,ABCDE,ABCD\n"
        "\n   - Alcance / Mecanismo: etiquetas de nodos (A=0, B=1, C=2, ...).\n"
    )

    # Candidato y estado inicial — comunes a todo el lote
    bits_candidato, n_candidatos = _pedir_candidato(n_nodos)
    _, estado = _pedir_estado(n_candidatos, bits_candidato, n_nodos)

    # K global para todo el lote
    print(f"\n Número máximo de particiones posibles: {n_candidatos}")
    k_input = input(
        f" Ingrese el valor de K para todo el lote [2-{n_candidatos}]\n"
        f"   O presione [ENTER] para evaluar TODAS las k posibles: "
    ).strip()

    if not k_input:
        k_val_global = None
        k_nombre = "All"
        print(f" Se evaluarán TODAS las k posibles en cada prueba.")
    else:
        try:
            k_val_global = int(k_input)
            if k_val_global < 2 or k_val_global > n_candidatos:
                print(f" Error: K debe estar entre 2 y {n_candidatos}.")
                sys.exit(1)
            k_nombre = str(k_val_global)
            print(f" K global = {k_val_global}")
        except ValueError:
            print(" Error: K debe ser un número entero.")
            sys.exit(1)

    # Seleccionar CSV de pruebas
    print("\n Seleccione el archivo CSV con las pruebas...")
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        ruta_csv_str = filedialog.askopenfilename(
            title="Seleccionar CSV de pruebas",
            filetypes=(("CSV files", "*.csv"), ("All files", "*.*"))
        )
        root.destroy()
    except Exception as e:
        print(f" No se pudo abrir el explorador: {e}")
        ruta_csv_str = input(" Ingrese la ruta del CSV de pruebas: ").strip()

    if not ruta_csv_str:
        print(" Ejecución cancelada: No se seleccionó CSV de pruebas.")
        sys.exit(1)

    ruta_csv = Path(ruta_csv_str)
    if not ruta_csv.exists():
        print(f" Error: No se encontró el archivo {ruta_csv}.")
        sys.exit(1)

    # Ruta de salida fija: results/block/results_kGeoMIP_k{k}_N{n}.xlsx
    results_block_dir = GEOMIP_ROOT / "results" / "block"
    results_block_dir.mkdir(parents=True, exist_ok=True)
    nombre_excel = f"results_kGeoMIP_k{k_nombre}_N{n_nodos}.xlsx"
    ruta_salida = results_block_dir / nombre_excel

    # Leer CSV
    with open(ruta_csv, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        pruebas = list(reader)

    if not pruebas:
        print(" Error: El CSV de pruebas está vacío.")
        sys.exit(1)

    def _get_col(row: dict, *nombres: str) -> str:
        for nombre in nombres:
            for key, val in row.items():
                if key.strip().lower() == nombre.lower():
                    return str(val).strip() if val is not None else ""
        return ""

    # Etiquetar cada prueba con su posición original en el CSV antes de reordenar.
    # El Excel se guardará en el orden original; la ejecución va de menor a mayor
    # complejidad para aprovechar los cachés calientes.
    for orig_idx, row in enumerate(pruebas):
        row["_orig_idx"] = orig_idx

    pruebas.sort(key=lambda row: (
        len(_get_col(row, "alcance o purview (t+1)", "alcance", "purview", "effect")),
        len(_get_col(row, "mecanismo(t)", "mecanismo", "mechanism", "cause")),
    ))
    print(" Pruebas reordenadas de menor a mayor complejidad (por nodos en alcance).")

    sep(f" Iniciando lote: {len(pruebas)} prueba(s) — '{ruta_csv.name}'")
    print(f"   Dataset TPM     : {ruta_tpm.name}")
    print(f"   Sist. Candidato : {bits_candidato}  ({n_candidatos} nodos)")
    print(f"   Estado Inicial  : {estado}")
    print(f"   K global        : {k_nombre}")
    print(f"   Salida          : {ruta_salida}\n")

    # Crear Manager y KGeoMIP una sola vez para todo el lote.
    # aplicar_estrategia limpia sus cachés internos al inicio de cada llamada,
    # por lo que reutilizar la instancia es seguro y evita recrear SafeLoggers
    # (6 mkdir + 4 FileHandler opens) en cada prueba.
    manager_inst = Manager(estado)
    kgeomip_inst = KGeoMIP(manager_inst)

    resultados = []

    def _silenciar_consola():
        """Silencia los StreamHandlers de consola de todos los loggers activos."""
        for nombre in list(logging.Logger.manager.loggerDict):
            log = logging.getLogger(nombre)
            for h in log.handlers:
                if type(h) is logging.StreamHandler:
                    h._nivel_previo = h.level
                    h.setLevel(999)

    def _restaurar_consola():
        """Restaura los StreamHandlers de consola al nivel original."""
        for nombre in list(logging.Logger.manager.loggerDict):
            log = logging.getLogger(nombre)
            for h in log.handlers:
                if hasattr(h, "_nivel_previo"):
                    h.setLevel(h._nivel_previo)
                    del h._nivel_previo

    total = len(pruebas)
    for idx, row in enumerate(pruebas, start=1):
        num_prueba   = _get_col(row, "#prueba", "prueba", "num", "n") or str(idx)
        alcance_raw  = _get_col(row, "alcance o purview (t+1)", "alcance", "purview", "effect")
        mecanism_raw = _get_col(row, "mecanismo(t)", "mecanismo", "mechanism", "cause")

        orig_idx     = row.get("_orig_idx", idx - 1)
        print(f" Prueba #{num_prueba:<4} ({idx}/{total})", end="", flush=True)

        if not alcance_raw or not mecanism_raw:
            msg = "Alcance o mecanismo vacíos en el CSV"
            print(f"  ✗ {msg}")
            resultados.append({
                "_orig_idx": orig_idx,
                "#prueba": num_prueba,
                "alcance_orig": alcance_raw,
                "mecanismo_orig": mecanism_raw,
                "error": msg,
            })
            continue

        # Convertir etiquetas de nodos a máscara binaria
        bits_alcance   = _letras_a_binario(alcance_raw, n_nodos)
        bits_mecanismo = _letras_a_binario(mecanism_raw, n_nodos)

        invalidos_alc = [i for i, (a, c) in enumerate(zip(bits_alcance, bits_candidato)) if a == "1" and c == "0"]
        if invalidos_alc:
            msg = f"Alcance activa nodos {invalidos_alc} fuera del candidato"
            print(f"  ✗ {msg}")
            resultados.append({"_orig_idx": orig_idx, "#prueba": num_prueba, "alcance_orig": alcance_raw, "mecanismo_orig": mecanism_raw, "error": msg})
            continue

        invalidos_mec = [i for i, (m, c) in enumerate(zip(bits_mecanismo, bits_candidato)) if m == "1" and c == "0"]
        if invalidos_mec:
            msg = f"Mecanismo activa nodos {invalidos_mec} fuera del candidato"
            print(f"  ✗ {msg}")
            resultados.append({"_orig_idx": orig_idx, "#prueba": num_prueba, "alcance_orig": alcance_raw, "mecanismo_orig": mecanism_raw, "error": msg})
            continue

        # Ejecución
        try:
            _silenciar_consola()
            try:
                solucion = kgeomip_inst.aplicar_estrategia(
                    condicion=bits_candidato,
                    alcance=bits_alcance,
                    mecanismo=bits_mecanismo,
                    tpm=tpm,
                    k=k_val_global,
                )
            finally:
                _restaurar_consola()

            tiempo_seg = float(
                getattr(solucion, "tiempo_ejecucion",
                        getattr(solucion, "tiempo_total", 0.0))
            )
            estrateg_str = getattr(solucion, "estrategia", "kgeomip")
            match_k = re.search(r'K=(\d+)', estrateg_str)
            k_optima = int(match_k.group(1)) if match_k else None

            particion_str = str(solucion.particion)
            tiempo_fmt = formatear_tiempo(tiempo_seg)

            print(f"  ✓  {tiempo_fmt}")

            resultados.append({
                "_orig_idx": orig_idx,
                "#prueba": num_prueba,
                "alcance_orig": alcance_raw,
                "mecanismo_orig": mecanism_raw,
                "k_optima": k_optima,
                "perdida_phi": float(solucion.perdida),
                "particion": particion_str,
                "tiempo_formateado": tiempo_fmt,
                "error": None,
            })

        except Exception as e:
            _restaurar_consola()
            msg = str(e)
            print(f"  ✗  Error: {msg}")
            resultados.append({
                "_orig_idx": orig_idx,
                "#prueba": num_prueba,
                "alcance_orig": alcance_raw,
                "mecanismo_orig": mecanism_raw,
                "error": msg,
            })

    # Restaurar orden original del CSV antes de guardar.
    resultados.sort(key=lambda r: r.get("_orig_idx", 0))

    # Guardar Excel
    _guardar_excel_bloque(ruta_salida, resultados)

    exitosas = sum(1 for r in resultados if r.get("error") is None)
    con_error = len(resultados) - exitosas

    sep(" Lote Completado")
    print(f"   Total pruebas   : {len(pruebas)}")
    print(f"   Exitosas        : {exitosas}")
    print(f"   Con error       : {con_error}")
    print(f"\n ✓ Resultados guardados en: {ruta_salida}")


# ── Punto de entrada ───────────────────────────────────────────────────────

def main():
    aplicacion.profiler_habilitado = False

    sep(" KGeoMIP — k-Partición de Mínima Información (Óptimo Global)")

    opcion = input(
        "\n ¿Qué modo desea ejecutar?\n"
        "   1. Manual     — ingresar una prueba de forma interactiva\n"
        "   2. Por bloque — cargar lote de pruebas desde un archivo CSV\n"
        " Seleccione (1 o 2): "
    ).strip()

    print()
    ruta_archivo, tpm, n_nodos = _seleccionar_tpm()
    print(f" TPM cargada: {ruta_archivo.name}  ({n_nodos} nodos)\n")

    if opcion == "2":
        modo_bloque(ruta_archivo, tpm, n_nodos)
    else:
        sep(" Ejecución Manual — KGeoMIP Óptimo Global")
        modo_manual(ruta_archivo, tpm, n_nodos)


if __name__ == '__main__':
    main()
